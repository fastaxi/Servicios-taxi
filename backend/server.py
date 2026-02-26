from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from jose import JWTError, jwt
from passlib.context import CryptContext
from datetime import datetime, timedelta

import subprocess

def get_git_sha() -> str:
    """Return short git SHA if available, else env GIT_SHA, else 'unknown'."""
    # Prefer real repo SHA when .git exists (Emergent preview), then fall back to env
    try:
        sha = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=str(pathlib.Path(__file__).resolve().parent.parent),
            stderr=subprocess.DEVNULL,
            text=True,
        ).strip()
        if sha:
            return sha
    except Exception:
        pass
    return os.environ.get("GIT_SHA", "unknown")
import pathlib
import os
import logging
import secrets
from pathlib import Path
from urllib.parse import urlparse
from pydantic import BaseModel, Field
from typing import List, Optional
from bson import ObjectId
import csv
import io
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

# Zona horaria de España
SPAIN_TZ = pytz.timezone('Europe/Madrid')

def get_spain_now():
    """Obtener la hora actual en España"""
    return datetime.now(SPAIN_TZ)

def parse_spanish_date_to_utc(fecha_str: str, hora_str: str = "00:00") -> Optional[datetime]:
    """
    Convierte fecha dd/mm/yyyy O yyyy-mm-dd + hora HH:mm (hora de España) a datetime UTC.
    Retorna None si el formato es inválido.
    """
    try:
        if not fecha_str:
            return None
        
        # Si ya es un datetime, devolverlo (para datos que ya fueron migrados)
        if isinstance(fecha_str, datetime):
            return fecha_str
        
        # Intentar parsear en diferentes formatos
        day, month, year = None, None, None
        
        # Formato español: dd/mm/yyyy
        if "/" in fecha_str:
            parts = fecha_str.split("/")
            if len(parts) == 3:
                day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
        # Formato ISO: yyyy-mm-dd
        elif "-" in fecha_str:
            parts = fecha_str.split("-")
            if len(parts) == 3:
                year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
        
        if day is None or month is None or year is None:
            return None
        
        # Parsear hora (default 00:00 si no se proporciona)
        hora_str = hora_str or "00:00"
        if isinstance(hora_str, datetime):
            hour, minute = hora_str.hour, hora_str.minute
        else:
            hora_parts = hora_str.split(":")
            hour = int(hora_parts[0]) if len(hora_parts) >= 1 else 0
            minute = int(hora_parts[1]) if len(hora_parts) >= 2 else 0
        
        # Crear datetime naive en hora de España
        local_dt = datetime(year, month, day, hour, minute, 0)
        # Localizar a España
        spain_dt = SPAIN_TZ.localize(local_dt)
        # Convertir a UTC
        utc_dt = spain_dt.astimezone(pytz.UTC).replace(tzinfo=None)
        return utc_dt
    except (ValueError, TypeError, IndexError, AttributeError) as e:
        logger.warning(f"Error parsing date '{fecha_str}' + '{hora_str}': {e}")
        return None

def get_date_range_utc(start_date: str, end_date: str) -> tuple:
    """
    Convierte rango de fechas dd/mm/yyyy (España) a rango UTC para queries.
    start_date 00:00:00 España -> UTC
    end_date 23:59:59 España -> UTC
    Retorna (start_utc, end_utc) o (None, None) si inválido.
    """
    start_utc = parse_spanish_date_to_utc(start_date, "00:00")
    end_utc = parse_spanish_date_to_utc(end_date, "23:59")
    if end_utc:
        # Añadir 59 segundos para incluir todo el último minuto
        end_utc = end_utc.replace(second=59, microsecond=999999)
    return (start_utc, end_utc)

# Simple in-memory cache
class SimpleCache:
    """Cache simple en memoria para datos consultados frecuentemente"""
    def __init__(self):
        self._cache = {}
        self._ttl = {}  # Time-to-live para cada key
    
    def get(self, key: str):
        """Obtener valor del cache si existe y no ha expirado"""
        if key in self._cache:
            # Verificar si no ha expirado (TTL de 5 minutos)
            if key in self._ttl and datetime.utcnow() < self._ttl[key]:
                return self._cache[key]
            else:
                # Expirado, eliminar
                self.delete(key)
        return None
    
    def set(self, key: str, value, ttl_minutes: int = 5):
        """Guardar valor en cache con TTL"""
        self._cache[key] = value
        self._ttl[key] = datetime.utcnow() + timedelta(minutes=ttl_minutes)
    
    def delete(self, key: str):
        """Eliminar valor del cache"""
        if key in self._cache:
            del self._cache[key]
        if key in self._ttl:
            del self._ttl[key]
    
    def clear(self):
        """Limpiar todo el cache"""
        self._cache.clear()
        self._ttl.clear()

# Instancia global del cache
cache = SimpleCache()

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env', override=False)

# MongoDB connection - Compatible con desarrollo y producción
mongo_url = os.getenv('MONGO_URL') or os.getenv('MONGODB_URI')
db_name = os.getenv('DB_NAME') or os.getenv('MONGODB_DB_NAME')

# Validar configuración de MongoDB
if not mongo_url:
    raise ValueError("MONGO_URL or MONGODB_URI must be set in environment variables")

if not db_name:
    # En producción, extraer el nombre de la base de datos de la URL si no está especificado
    if 'mongodb+srv://' in mongo_url or 'mongodb://' in mongo_url:
        # Intentar extraer el nombre de la BD de la URL
        parsed = urlparse(mongo_url)
        if parsed.path and len(parsed.path) > 1:
            db_name = parsed.path[1:].split('?')[0]
            print(f"[STARTUP] Database name extracted from URL: {db_name}")
    
    if not db_name:
        raise ValueError("DB_NAME or MONGODB_DB_NAME must be set in environment variables")

# Log configuration for debugging
print(f"[STARTUP] Connecting to MongoDB...")
print(f"[STARTUP] Database: {db_name}")
print(f"[STARTUP] MongoDB URL type: {'Atlas' if 'mongodb+srv://' in mongo_url else 'Local'}")

try:
    # Configuración optimizada para MongoDB Atlas y local
    client = AsyncIOMotorClient(
        mongo_url,
        serverSelectionTimeoutMS=10000,
        connectTimeoutMS=10000,
        socketTimeoutMS=10000,
        maxPoolSize=50,
        minPoolSize=10
    )
    db = client[db_name]
    print("[STARTUP] MongoDB connection initialized successfully")
except Exception as e:
    print(f"[STARTUP ERROR] Failed to connect to MongoDB: {e}")
    raise

# Configure logging FIRST (before any log calls)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Security
SECRET_KEY = os.environ.get("SECRET_KEY")
ENV = os.environ.get("ENV", "development").lower()

# Límite para batch queries de servicios (configurable vía env)
MAX_BATCH_SERVICES = int(os.environ.get("MAX_BATCH_SERVICES", "10000"))

# ID de la organización Taxitur (para reglas específicas de origen parada/lagos)
TAXITUR_ORG_ID = os.environ.get("TAXITUR_ORG_ID", "69484bec187c3bc2b0fdb8f4")

# Startup logging (sin revelar longitud de secretos)
logger.info("=" * 60)
logger.info("TAXIFAST API STARTING")
logger.info(f"  ENV: {ENV}")
logger.info(f"  MAX_BATCH_SERVICES: {MAX_BATCH_SERVICES}")
logger.info(f"  TAXITUR_ORG_ID: {'configured' if TAXITUR_ORG_ID else 'missing'}")

if not SECRET_KEY:
    if ENV == "production":
        # FAIL-FAST en producción: no arrancar sin SECRET_KEY
        logger.critical("SECRET_KEY not set in production. Refusing to start.")
        raise RuntimeError(
            "SECRET_KEY not set in production. Set SECRET_KEY in environment variables."
        )
    # Development / staging: clave temporal aceptable
    SECRET_KEY = secrets.token_hex(32)
    logger.warning("SECRET_KEY: temporary (sessions lost on restart)")
else:
    logger.info("  SECRET_KEY: configured")  # Sin revelar longitud

logger.info("=" * 60)

ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30 * 24 * 60  # 30 days

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Create the main app - TaxiFast Multi-tenant SaaS Platform
app = FastAPI(title="TaxiFast API", version="2.0.0", description="Multi-tenant taxi management SaaS platform")
api_router = APIRouter(prefix="/api")

# ==========================================
# MIDDLEWARE: Request Logging & Metrics
# ==========================================
import time
import uuid
from collections import defaultdict
from threading import Lock

# Umbrales de latencia por tipo de endpoint (ms)
SLOW_THRESHOLD_DEFAULT = 1000  # 1 segundo para endpoints normales
SLOW_THRESHOLD_EXPORT = 5000   # 5 segundos para exports

# Métricas en memoria para monitoreo
class MetricsCollector:
    def __init__(self):
        self._lock = Lock()
        self._error_counts = defaultdict(int)  # {endpoint: count}
        self._slow_requests = []  # Lista de requests lentos recientes
        self._total_requests = 0
        self._total_errors_5xx = 0
        self._total_errors_4xx = 0
        self._start_time = datetime.utcnow()
    
    def record_request(self, method: str, path: str, status: int, duration_ms: float):
        with self._lock:
            self._total_requests += 1
            
            if status >= 500:
                self._total_errors_5xx += 1
                self._error_counts[f"{method} {path}"] += 1
            elif status >= 400:
                self._total_errors_4xx += 1
            
            # Guardar requests lentos (>2s) - máximo 50 recientes
            if duration_ms > 2000:
                self._slow_requests.append({
                    "time": datetime.utcnow().isoformat(),
                    "method": method,
                    "path": path,
                    "duration_ms": round(duration_ms),
                    "status": status
                })
                if len(self._slow_requests) > 50:
                    self._slow_requests.pop(0)
    
    def get_metrics(self):
        with self._lock:
            uptime = (datetime.utcnow() - self._start_time).total_seconds()
            return {
                "uptime_seconds": round(uptime),
                "total_requests": self._total_requests,
                "total_5xx_errors": self._total_errors_5xx,
                "total_4xx_errors": self._total_errors_4xx,
                "error_rate_5xx": round(self._total_errors_5xx / max(self._total_requests, 1) * 100, 2),
                "top_error_endpoints": dict(sorted(self._error_counts.items(), key=lambda x: -x[1])[:10]),
                "recent_slow_requests": self._slow_requests[-10:],
                "alerts": self._check_alerts()
            }
    
    def _check_alerts(self):
        alerts = []
        # Alerta si error rate > 5%
        if self._total_requests > 100:
            error_rate = self._total_errors_5xx / self._total_requests * 100
            if error_rate > 5:
                alerts.append(f"HIGH_ERROR_RATE: {error_rate:.1f}% de requests con 5xx")
        
        # Alerta si más de 10 errores 5xx en la sesión
        if self._total_errors_5xx > 10:
            alerts.append(f"MANY_5XX_ERRORS: {self._total_errors_5xx} errores 5xx desde el arranque")
        
        return alerts

# Instancia global de métricas
metrics = MetricsCollector()

@app.middleware("http")
async def log_requests(request, call_next):
    """Log estructurado de cada request con tiempo de respuesta y Request ID"""
    start_time = time.time()
    
    # Generar Request ID único para trazabilidad
    request_id = str(uuid.uuid4())[:8]  # 8 chars suficiente para debugging
    
    # Procesar request
    response = await call_next(request)
    
    # Calcular tiempo
    process_time = (time.time() - start_time) * 1000  # ms
    
    # Añadir headers
    response.headers["X-Request-Id"] = request_id
    response.headers["X-Process-Time"] = f"{process_time:.0f}ms"
    
    # Log estructurado (solo para /api, excluir health checks)
    path = request.url.path
    if path.startswith("/api") and path not in ["/api/health", "/", "/api/metrics"]:
        status = response.status_code
        method = request.method
        
        # Registrar métricas
        metrics.record_request(method, path, status, process_time)
        
        # Determinar umbral de latencia según endpoint
        is_export = "/export/" in path
        slow_threshold = SLOW_THRESHOLD_EXPORT if is_export else SLOW_THRESHOLD_DEFAULT
        
        # Nivel de log según status code (reducir ruido en 4xx esperables)
        log_msg = f"[{request_id}] [{method}] {path} -> {status} ({process_time:.0f}ms)"
        
        if status >= 500:
            # Errores de servidor: siempre ERROR
            logger.error(log_msg)
        elif status == 429:
            # Rate limit: WARNING (importante)
            logger.warning(log_msg)
        elif status >= 400:
            # 4xx esperables (400/401/403/404/422): INFO para reducir ruido
            logger.info(log_msg)
        elif process_time > slow_threshold:
            # Request lento: WARNING con indicador
            logger.warning(f"{log_msg} SLOW")
        else:
            # Normal: INFO
            logger.info(log_msg)
    
    return response

# Root health check endpoint for deployment systems
@app.get("/")
async def root_health_check():
    """Health check endpoint for deployment verification"""
    GIT_SHA = get_git_sha()
    return {
        "status": "healthy",
        "service": "taxifast-api",
        "version": "1.0.0",
        "git_sha": GIT_SHA,
        "timestamp": datetime.utcnow().isoformat()
    }

# API health check endpoint
@app.get("/health")
async def health_check():
    """Detailed health check with database connectivity"""
    try:
        # Test database connection
        await db.command("ping")
        db_status = "connected"
    except Exception as e:
        logger.error(f"Database health check failed: {e}")
        db_status = "disconnected"
    
    GIT_SHA = get_git_sha()
    return {
        "status": "healthy" if db_status == "connected" else "degraded",
        "database": db_status,
        "git_sha": GIT_SHA,
        "timestamp": datetime.utcnow().isoformat()
    }

# Helper function for ObjectId
class PyObjectId(ObjectId):
    @classmethod
    def __get_validators__(cls):
        yield cls.validate

    @classmethod
    def validate(cls, v):
        if not ObjectId.is_valid(v):
            raise ValueError("Invalid ObjectId")
        return ObjectId(v)

    @classmethod
    def __get_pydantic_json_schema__(cls, field_schema):
        field_schema.update(type="string")

# ==========================================
# ORGANIZATION MODELS (Multi-tenant SaaS)
# ==========================================
class OrganizationBase(BaseModel):
    nombre: str  # Nombre de la empresa de taxis (ej: "TaxiFast", "Radio Taxi Madrid")
    slug: Optional[str] = None  # URL-friendly identifier (auto-generated if not provided)
    cif: Optional[str] = ""  # CIF/NIF de la empresa
    direccion: Optional[str] = ""
    codigo_postal: Optional[str] = ""
    localidad: Optional[str] = ""
    provincia: Optional[str] = ""
    telefono: Optional[str] = ""
    email: Optional[str] = ""
    web: Optional[str] = ""
    logo_base64: Optional[str] = None  # Logo de la empresa en base64
    color_primario: Optional[str] = "#0066CC"  # Color principal de la marca
    color_secundario: Optional[str] = "#FFD700"  # Color secundario
    notas: Optional[str] = ""
    activa: bool = True  # Si la organización está activa
    features: Optional[dict] = None  # Feature flags: {"taxitur_origen": true/false, ...}
    settings: Optional[dict] = None  # Tenant settings: branding, footer, etc.

# Whitelist de feature flags permitidas
ALLOWED_FEATURE_KEYS = {
    "taxitur_origen",  # Habilitar campo origen Parada/Lagos
}

# Whitelist de keys permitidas para settings (evitar que sea un basurero)
ALLOWED_SETTINGS_KEYS = {
    "display_name",      # Nombre a mostrar en UI
    "logo_url",          # URL del logo (alternativa a base64)
    "footer_name",       # Nombre para el footer
    "footer_cif",        # CIF para el footer
    "footer_extra",      # Texto extra para footer
    "primary_color",     # Color primario (override)
    "support_email",     # Email de soporte
    "support_phone",     # Telefono de soporte
}

class OrganizationCreate(OrganizationBase):
    pass

class OrganizationUpdate(BaseModel):
    nombre: Optional[str] = None
    slug: Optional[str] = None
    cif: Optional[str] = None
    direccion: Optional[str] = None
    codigo_postal: Optional[str] = None
    localidad: Optional[str] = None
    provincia: Optional[str] = None
    telefono: Optional[str] = None
    email: Optional[str] = None
    web: Optional[str] = None
    logo_base64: Optional[str] = None
    color_primario: Optional[str] = None
    color_secundario: Optional[str] = None
    notas: Optional[str] = None
    activa: Optional[bool] = None
    features: Optional[dict] = None  # Feature flags
    settings: Optional[dict] = None  # Tenant settings

class OrganizationResponse(OrganizationBase):
    id: str
    created_at: datetime
    updated_at: Optional[datetime] = None
    features: Optional[dict] = None  # Feature flags
    settings: Optional[dict] = None  # Tenant settings
    # Estadísticas calculadas
    total_taxistas: Optional[int] = 0
    total_vehiculos: Optional[int] = 0
    total_clientes: Optional[int] = 0

    class Config:
        from_attributes = True

# ==========================================
# USER MODELS (Multi-tenant aware)
# ==========================================
class UserBase(BaseModel):
    username: str
    nombre: str
    role: str = "taxista"  # superadmin, admin or taxista
    licencia: Optional[str] = None
    vehiculo_id: Optional[str] = None
    vehiculo_matricula: Optional[str] = None  # Para mostrar sin hacer joins
    organization_id: Optional[str] = None  # ID de la organización (null para superadmin)

class UserCreate(UserBase):
    password: str

class UserResponse(UserBase):
    id: str
    created_at: datetime
    organization_nombre: Optional[str] = None  # Nombre de la organización para display

    class Config:
        from_attributes = True

class UserLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

# ==========================================
# COMPANY/CLIENT MODELS (Multi-tenant aware)
# ==========================================
class CompanyBase(BaseModel):
    nombre: str
    cif: Optional[str] = ""
    direccion: Optional[str] = ""
    codigo_postal: Optional[str] = ""
    localidad: Optional[str] = ""
    provincia: Optional[str] = ""
    telefono: Optional[str] = ""
    email: Optional[str] = ""
    numero_cliente: Optional[str] = None
    contacto: Optional[str] = ""
    fecha_alta: Optional[str] = None  # formato dd/mm/yyyy
    notas: Optional[str] = ""
    # organization_id se asigna automáticamente del usuario actual

class CompanyCreate(CompanyBase):
    pass

class CompanyResponse(CompanyBase):
    id: str
    created_at: datetime
    organization_id: Optional[str] = None

    class Config:
        from_attributes = True

class ServiceBase(BaseModel):
    fecha: str
    hora: str
    origen: str
    destino: str
    importe: float  # IVA 10% incluido
    importe_espera: float  # Importe de espera en euros
    importe_total: Optional[float] = None  # Se calcula automaticamente
    kilometros: Optional[float] = None  # Ahora opcional (B)
    tipo: str  # "empresa" or "particular"
    empresa_id: Optional[str] = None
    empresa_nombre: Optional[str] = None
    turno_id: Optional[str] = None  # ID del turno asociado
    cobrado: bool = False
    facturar: bool = False
    # Nuevos campos funcionales PR1
    metodo_pago: Optional[str] = "efectivo"  # "efectivo" | "tpv" | null (D)
    origen_taxitur: Optional[str] = None  # "parada" | "lagos" - Solo para org Taxitur (E)
    vehiculo_id: Optional[str] = None  # Vehiculo usado en el servicio (A)
    vehiculo_matricula: Optional[str] = None  # Denormalizado para exports (A)
    vehiculo_cambiado: Optional[bool] = False  # True si uso vehiculo distinto al default (A)
    km_inicio_vehiculo: Optional[int] = None  # Obligatorio si vehiculo_cambiado=True (A)
    km_fin_vehiculo: Optional[int] = None  # Obligatorio si vehiculo_cambiado=True (A)
    # Idempotencia (Paso 5A)
    client_uuid: Optional[str] = None  # UUID generado por cliente para evitar duplicados

class ServiceCreate(ServiceBase):
    pass

class ServiceResponse(BaseModel):
    id: str
    taxista_id: str
    taxista_nombre: str
    fecha: str
    hora: str
    origen: str
    destino: str
    importe: float
    importe_espera: Optional[float] = 0  # Opcional para compatibilidad con datos antiguos
    importe_total: Optional[float] = 0  # Opcional para compatibilidad con datos antiguos
    kilometros: Optional[float] = None  # Ahora opcional (B)
    tipo: str
    empresa_id: Optional[str] = None
    empresa_nombre: Optional[str] = None
    turno_id: Optional[str] = None
    cobrado: Optional[bool] = False
    facturar: Optional[bool] = False
    created_at: datetime
    synced: bool = True
    organization_id: Optional[str] = None  # Multi-tenant support
    # Nuevos campos funcionales PR1
    metodo_pago: Optional[str] = None  # "efectivo" | "tpv" (D)
    origen_taxitur: Optional[str] = None  # "parada" | "lagos" - Solo para org Taxitur (E)
    vehiculo_id: Optional[str] = None  # Vehiculo usado en el servicio (A)
    vehiculo_matricula: Optional[str] = None  # Denormalizado para exports (A)
    vehiculo_cambiado: Optional[bool] = False  # True si uso vehiculo distinto al default (A)
    km_inicio_vehiculo: Optional[int] = None  # Obligatorio si vehiculo_cambiado=True (A)
    km_fin_vehiculo: Optional[int] = None  # Obligatorio si vehiculo_cambiado=True (A)
    # Idempotencia (Paso 5A)
    client_uuid: Optional[str] = None  # UUID del cliente para idempotencia

    class Config:
        from_attributes = True

class ConfigBase(BaseModel):
    nombre_radio_taxi: str
    telefono: str
    web: str
    direccion: Optional[str] = ""
    email: Optional[str] = ""
    logo_base64: Optional[str] = None

class ConfigResponse(ConfigBase):
    id: str
    updated_at: datetime

    class Config:
        from_attributes = True

class ServiceSync(BaseModel):
    services: List[ServiceBase]

# ==========================================
# VEHICULO MODELS (Multi-tenant aware)
# ==========================================
class VehiculoBase(BaseModel):
    matricula: str
    plazas: int
    marca: str
    modelo: str
    km_iniciales: int
    fecha_compra: str  # formato dd/mm/yyyy
    activo: bool = True
    # organization_id se asigna automáticamente del usuario actual

class VehiculoCreate(VehiculoBase):
    pass

class VehiculoResponse(BaseModel):
    id: str
    matricula: str
    plazas: Optional[int] = 4  # Default value for compatibility
    marca: Optional[str] = ""
    modelo: Optional[str] = ""
    km_iniciales: Optional[int] = 0  # Default value for compatibility
    fecha_compra: Optional[str] = ""  # Default value for compatibility
    activo: Optional[bool] = True
    organization_id: Optional[str] = None

    class Config:
        from_attributes = True

# ==========================================
# TURNO MODELS (Multi-tenant aware)
# ==========================================
class TurnoBase(BaseModel):
    taxista_id: str
    taxista_nombre: str
    vehiculo_id: str
    vehiculo_matricula: str
    fecha_inicio: str  # formato dd/mm/yyyy
    hora_inicio: str   # formato HH:mm
    km_inicio: int
    fecha_fin: Optional[str] = None
    hora_fin: Optional[str] = None
    km_fin: Optional[int] = None
    cerrado: bool = False
    liquidado: bool = False
    # organization_id se asigna automáticamente del usuario actual

# Modelo de Combustible para turnos (F)
class CombustibleData(BaseModel):
    repostado: bool = False
    litros: Optional[float] = None  # Obligatorio si repostado=True, debe ser > 0
    vehiculo_id: Optional[str] = None  # Obligatorio si repostado=True
    vehiculo_matricula: Optional[str] = None  # Denormalizado para exports
    km_vehiculo: Optional[int] = None  # Obligatorio si repostado=True, >= 0
    timestamp: Optional[datetime] = None  # Server time cuando se registra
    registrado_por_user_id: Optional[str] = None  # Para trazabilidad

class TurnoCreate(BaseModel):
    taxista_id: str
    taxista_nombre: str
    vehiculo_id: str
    vehiculo_matricula: str
    fecha_inicio: str
    hora_inicio: str  # Se ignorará y se usará hora del servidor (C)
    km_inicio: int

class TurnoUpdate(BaseModel):
    fecha_inicio: Optional[str] = None
    hora_inicio: Optional[str] = None
    km_inicio: Optional[int] = None
    fecha_fin: Optional[str] = None
    hora_fin: Optional[str] = None
    km_fin: Optional[int] = None
    cerrado: Optional[bool] = None
    liquidado: Optional[bool] = None

class TurnoFinalizarUpdate(BaseModel):
    fecha_fin: str
    hora_fin: str  # Se ignorará y se usará hora del servidor (C)
    km_fin: int
    cerrado: bool = True

class CombustibleUpdate(BaseModel):
    """Modelo para actualizar combustible en un turno (F)"""
    repostado: bool
    litros: Optional[float] = None
    vehiculo_id: Optional[str] = None
    km_vehiculo: Optional[int] = None

class TurnoResponse(TurnoBase):
    id: str
    created_at: datetime
    organization_id: Optional[str] = None
    # Totales calculados
    total_importe_clientes: Optional[float] = 0
    total_importe_particulares: Optional[float] = 0
    total_kilometros: Optional[float] = 0
    cantidad_servicios: Optional[int] = 0
    # Combustible (F)
    combustible: Optional[CombustibleData] = None

    class Config:
        from_attributes = True

# ==========================================
# AUTH HELPERS (Multi-tenant support)
# ==========================================
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def generate_slug(nombre: str) -> str:
    """Genera un slug URL-friendly a partir del nombre"""
    import re
    # Convertir a minúsculas y reemplazar espacios por guiones
    slug = nombre.lower().strip()
    # Reemplazar caracteres especiales españoles
    replacements = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ñ': 'n', 'ü': 'u'}
    for old, new in replacements.items():
        slug = slug.replace(old, new)
    # Solo permitir alfanuméricos y guiones
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'[\s_]+', '-', slug)
    slug = re.sub(r'-+', '-', slug)
    return slug.strip('-')

def _get_object_id_or_400(id_str: str, field_name: str = "ID") -> ObjectId:
    """Convierte string a ObjectId o lanza 400 si es inválido"""
    try:
        return ObjectId(id_str)
    except Exception:
        raise HTTPException(status_code=400, detail=f"{field_name} tiene formato inválido")

async def _get_taxista_or_400(taxista_id: str, org_filter: dict, db_instance) -> dict:
    """Valida que el taxista existe y pertenece al scope. Retorna el taxista o lanza 400."""
    try:
        oid = ObjectId(taxista_id)
    except Exception:
        raise HTTPException(status_code=400, detail="taxista_id tiene formato inválido")
    
    taxista = await db_instance.users.find_one({"_id": oid, "role": "taxista", **org_filter})
    if not taxista:
        raise HTTPException(status_code=400, detail="El taxista especificado no existe o no pertenece a esta organización")
    return taxista

async def _get_company_or_400(empresa_id: str, org_filter: dict, db_instance) -> dict:
    """Valida que la empresa existe y pertenece al scope. Retorna la empresa o lanza 400."""
    try:
        oid = ObjectId(empresa_id)
    except Exception:
        raise HTTPException(status_code=400, detail="empresa_id tiene formato inválido")
    
    empresa = await db_instance.companies.find_one({"_id": oid, **org_filter})
    if not empresa:
        raise HTTPException(status_code=400, detail="La empresa especificada no existe o no pertenece a esta organización")
    return empresa

async def get_empresa_or_400(empresa_id: str, current_user: dict):
    """Wrapper conveniente para validar empresa_id con current_user."""
    org_filter = await get_org_filter(current_user)
    return await _get_company_or_400(empresa_id, org_filter, db)

async def _get_turno_or_400(turno_id: str, org_filter: dict, db_instance, taxista_id: str = None) -> dict:
    """Valida que el turno existe y pertenece al scope. Si taxista_id, además verifica que sea suyo."""
    try:
        oid = ObjectId(turno_id)
    except Exception:
        raise HTTPException(status_code=400, detail="turno_id tiene formato inválido")
    
    query = {"_id": oid, **org_filter}
    if taxista_id:
        query["taxista_id"] = taxista_id
    
    turno = await db_instance.turnos.find_one(query)
    if not turno:
        raise HTTPException(status_code=400, detail="El turno especificado no existe, no pertenece a esta organización, o no es tuyo")
    return turno

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"username": username})
    if user is None:
        raise credentials_exception
    
    # Verificar que la organización del usuario esté activa (excepto superadmin)
    if user.get("role") != "superadmin" and user.get("organization_id"):
        org = await db.organizations.find_one({"_id": ObjectId(user["organization_id"])})
        if org and not org.get("activa", True):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Tu organización está desactivada. Contacta con el administrador."
            )
    
    return user

async def get_current_admin(current_user: dict = Depends(get_current_user)):
    """Permite acceso a admin y superadmin"""
    if current_user.get("role") not in ["admin", "superadmin"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    return current_user

async def get_current_superadmin(current_user: dict = Depends(get_current_user)):
    """Solo permite acceso a superadmin"""
    if current_user.get("role") != "superadmin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Super admin permissions required"
        )
    return current_user

# ==========================================
# METRICS ENDPOINT (para monitoreo y alertas)
# ==========================================
@api_router.get("/metrics")
async def get_api_metrics(current_user: dict = Depends(get_current_user)):
    """
    Endpoint de métricas para monitoreo.
    Solo accesible por admin/superadmin.
    
    Retorna:
    - Uptime del servidor
    - Total de requests
    - Errores 5xx y 4xx
    - Tasa de errores
    - Top endpoints con errores
    - Requests lentos recientes
    - Alertas activas
    """
    # Solo admin o superadmin pueden ver métricas
    if current_user.get("role") not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver métricas")
    
    return {
        "timestamp": datetime.utcnow().isoformat(),
        "env": ENV,
        **metrics.get_metrics()
    }

def get_user_organization_id(user: dict) -> Optional[str]:
    """Obtiene el organization_id del usuario actual"""
    return user.get("organization_id")

def is_superadmin(user: dict) -> bool:
    """Verifica si el usuario es superadmin"""
    return user.get("role") == "superadmin"

async def get_org_filter(user: dict) -> dict:
    """
    Retorna el filtro de organización para queries.
    Superadmin ve todo, otros usuarios solo ven datos de su organización.
    SEGURIDAD: Si no es superadmin y no tiene organización, lanza 403.
    """
    if is_superadmin(user):
        return {}  # Sin filtro, ve todo
    
    org_id = get_user_organization_id(user)
    if org_id:
        return {"organization_id": org_id}
    
    # SEGURIDAD P0: Usuario sin organización no puede acceder a datos
    raise HTTPException(
        status_code=403, 
        detail="Usuario sin organización asignada. Contacte al administrador."
    )

# ==========================================
# HEALTH CHECK ENDPOINT (P1)
# ==========================================
@api_router.get("/health")
async def health_check():
    """
    Health check endpoint para monitoreo.
    Verifica conexión a la base de datos.
    """
    try:
        # Verificar conexión a MongoDB
        await db.command("ping")
        
        # Contar documentos básicos para verificar acceso
        users_count = await db.users.count_documents({})
        orgs_count = await db.organizations.count_documents({})
        
        return {
            "status": "healthy",
            "database": "connected",
            "timestamp": datetime.utcnow().isoformat(),
            "stats": {
                "users": users_count,
                "organizations": orgs_count
            }
        }
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        raise HTTPException(
            status_code=503,
            detail=f"Service unhealthy: {str(e)}"
        )

# ==========================================
# AUTH ENDPOINTS
# ==========================================
@api_router.post("/auth/login", response_model=Token)
async def login(user_login: UserLogin):
    user = await db.users.find_one({"username": user_login.username})
    if not user or not verify_password(user_login.password, user["password"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password"
        )
    
    # Obtener nombre de la organización si existe
    org_nombre = None
    if user.get("organization_id"):
        org = await db.organizations.find_one({"_id": ObjectId(user["organization_id"])})
        if org:
            org_nombre = org.get("nombre")
    
    access_token = create_access_token(data={"sub": user["username"]})
    user_response = UserResponse(
        id=str(user["_id"]),
        username=user["username"],
        nombre=user["nombre"],
        role=user["role"],
        organization_id=user.get("organization_id"),
        organization_nombre=org_nombre,
        created_at=user["created_at"]
    )
    return Token(access_token=access_token, token_type="bearer", user=user_response)

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    # Obtener nombre de la organización si existe
    org_nombre = None
    if current_user.get("organization_id"):
        org = await db.organizations.find_one({"_id": ObjectId(current_user["organization_id"])})
        if org:
            org_nombre = org.get("nombre")
    
    return UserResponse(
        id=str(current_user["_id"]),
        username=current_user["username"],
        nombre=current_user["nombre"],
        role=current_user["role"],
        organization_id=current_user.get("organization_id"),
        organization_nombre=org_nombre,
        licencia=current_user.get("licencia"),
        vehiculo_id=current_user.get("vehiculo_id"),
        vehiculo_matricula=current_user.get("vehiculo_matricula"),
        created_at=current_user["created_at"]
    )

# ==========================================
# ORGANIZATION BRANDING (For mobile app)
# ==========================================
@api_router.get("/my-organization")
async def get_my_organization(current_user: dict = Depends(get_current_user)):
    """Obtener información de branding de la organización del usuario actual (para app móvil)"""
    org_id = current_user.get("organization_id")
    
    if not org_id:
        # Usuario sin organización (legacy o superadmin) - devolver branding por defecto
        return {
            "id": None,
            "nombre": "TaxiFast",
            "slug": "taxifast",
            "logo_base64": None,
            "color_primario": "#0066CC",
            "color_secundario": "#FFD700",
            "telefono": "",
            "email": "",
            "web": "www.taxifast.com",
            "direccion": "",
            "localidad": "",
            "provincia": "",
            "features": {},  # Sin features activos por defecto
            "settings": {},  # Sin settings por defecto
        }
    
    org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    if not org:
        raise HTTPException(status_code=404, detail="Organizacion no encontrada")
    
    return {
        "id": str(org["_id"]),
        "nombre": org.get("nombre", "TaxiFast"),
        "slug": org.get("slug", ""),
        "logo_base64": org.get("logo_base64"),
        "color_primario": org.get("color_primario", "#0066CC"),
        "color_secundario": org.get("color_secundario", "#FFD700"),
        "telefono": org.get("telefono", ""),
        "email": org.get("email", ""),
        "web": org.get("web", ""),
        "direccion": org.get("direccion", ""),
        "localidad": org.get("localidad", ""),
        "provincia": org.get("provincia", ""),
        "cif": org.get("cif", ""),
        "features": org.get("features", {}),  # Feature flags de la organización
        "settings": org.get("settings", {}),  # Tenant settings
    }

def validate_settings(settings: dict) -> dict:
    """Validar y sanitizar settings según whitelist"""
    validated = {}
    invalid_keys = []
    
    for key, value in settings.items():
        if key not in ALLOWED_SETTINGS_KEYS:
            invalid_keys.append(key)
            continue
        
        # Validar tipos: solo strings y booleanos, max 500 chars para strings
        if isinstance(value, str):
            if len(value) > 500:
                raise HTTPException(
                    status_code=400, 
                    detail=f"El valor de '{key}' excede 500 caracteres"
                )
            validated[key] = value
        elif isinstance(value, bool):
            validated[key] = value
        elif value is None:
            validated[key] = None
        else:
            raise HTTPException(
                status_code=400, 
                detail=f"El valor de '{key}' debe ser string, boolean o null"
            )
    
    if invalid_keys:
        raise HTTPException(
            status_code=400, 
            detail=f"Keys no permitidas en settings: {', '.join(invalid_keys)}. Permitidas: {', '.join(ALLOWED_SETTINGS_KEYS)}"
        )
    
    return validated

@api_router.put("/my-organization/settings")
async def update_my_organization_settings(
    settings_data: dict,
    current_user: dict = Depends(get_current_admin)
):
    """Actualizar settings de la organización del admin actual (Admin o Superadmin)"""
    org_id = current_user.get("organization_id")
    
    if not org_id:
        raise HTTPException(
            status_code=400, 
            detail="Usuario sin organizacion asociada. No puede actualizar settings."
        )
    
    # Validar y sanitizar settings
    validated_settings = validate_settings(settings_data)
    
    # Obtener settings actuales y mergear
    org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    if not org:
        raise HTTPException(status_code=404, detail="Organizacion no encontrada")
    
    current_settings = org.get("settings", {})
    # Merge: los nuevos valores sobreescriben los existentes
    merged_settings = {**current_settings, **validated_settings}
    
    # Actualizar en BD
    await db.organizations.update_one(
        {"_id": ObjectId(org_id)},
        {"$set": {"settings": merged_settings, "updated_at": datetime.utcnow()}}
    )
    
    # Devolver org actualizada
    updated_org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    return {
        "id": str(updated_org["_id"]),
        "nombre": updated_org.get("nombre"),
        "settings": updated_org.get("settings", {}),
        "features": updated_org.get("features", {}),
        "message": "Settings actualizados correctamente"
    }

@api_router.put("/superadmin/organizations/{org_id}/settings")
async def superadmin_update_org_settings(
    org_id: str,
    settings_data: dict,
    current_user: dict = Depends(get_current_superadmin)
):
    """Actualizar settings de cualquier organización (solo Superadmin)"""
    # Verificar que la org existe
    try:
        org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de organizacion invalido")
    
    if not org:
        raise HTTPException(status_code=404, detail="Organizacion no encontrada")
    
    # Validar y sanitizar settings
    validated_settings = validate_settings(settings_data)
    
    # Obtener settings actuales y mergear
    current_settings = org.get("settings", {})
    merged_settings = {**current_settings, **validated_settings}
    
    # Actualizar en BD
    await db.organizations.update_one(
        {"_id": ObjectId(org_id)},
        {"$set": {"settings": merged_settings, "updated_at": datetime.utcnow()}}
    )
    
    # Devolver org actualizada
    updated_org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    return {
        "id": str(updated_org["_id"]),
        "nombre": updated_org.get("nombre"),
        "settings": updated_org.get("settings", {}),
        "features": updated_org.get("features", {}),
        "message": "Settings actualizados correctamente"
    }


@api_router.put("/superadmin/organizations/{org_id}/features")
async def superadmin_update_org_features(
    org_id: str,
    body: dict,
    current_user: dict = Depends(get_current_superadmin)
):
    """Actualizar feature flags de una organizacion (solo Superadmin).
    Body: { "features": { "taxitur_origen": true/false } }
    Solo se permiten keys de ALLOWED_FEATURE_KEYS y valores boolean.
    Merge: actualiza solo las keys enviadas sin borrar las existentes.
    """
    features_input = body.get("features")
    if not isinstance(features_input, dict) or len(features_input) == 0:
        raise HTTPException(status_code=400, detail="Se requiere un objeto 'features' con al menos una key")

    # Validar keys y tipos
    for key, value in features_input.items():
        if key not in ALLOWED_FEATURE_KEYS:
            raise HTTPException(
                status_code=400,
                detail=f"Feature flag '{key}' no permitida. Permitidas: {sorted(ALLOWED_FEATURE_KEYS)}"
            )
        if not isinstance(value, bool):
            raise HTTPException(
                status_code=400,
                detail=f"El valor de '{key}' debe ser boolean (true/false), recibido: {type(value).__name__}"
            )

    # Verificar que la org existe
    try:
        org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="ID de organizacion invalido")

    if not org:
        raise HTTPException(status_code=404, detail="Organizacion no encontrada")

    # Merge con features actuales
    current_features = org.get("features") or {}
    merged_features = {**current_features, **features_input}

    await db.organizations.update_one(
        {"_id": ObjectId(org_id)},
        {"$set": {"features": merged_features, "updated_at": datetime.utcnow()}}
    )

    updated_org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    return {
        "id": str(updated_org["_id"]),
        "nombre": updated_org.get("nombre"),
        "features": updated_org.get("features", {}),
        "settings": updated_org.get("settings", {}),
        "message": "Features actualizados correctamente"
    }


# ==========================================
# ORGANIZATION ENDPOINTS (Superadmin only)
# ==========================================
@api_router.post("/organizations", response_model=OrganizationResponse)
async def create_organization(org: OrganizationCreate, current_user: dict = Depends(get_current_superadmin)):
    """Crear nueva organización (solo superadmin)"""
    org_dict = org.dict()
    
    # Generar slug si no se proporciona
    if not org_dict.get("slug"):
        org_dict["slug"] = generate_slug(org_dict["nombre"])
    
    # Verificar slug único
    existing = await db.organizations.find_one({"slug": org_dict["slug"]})
    if existing:
        # Agregar número al slug
        count = await db.organizations.count_documents({"slug": {"$regex": f"^{org_dict['slug']}"}})
        org_dict["slug"] = f"{org_dict['slug']}-{count + 1}"
    
    org_dict["created_at"] = datetime.utcnow()
    org_dict["updated_at"] = datetime.utcnow()
    
    result = await db.organizations.insert_one(org_dict)
    created_org = await db.organizations.find_one({"_id": result.inserted_id})
    
    return OrganizationResponse(
        id=str(created_org["_id"]),
        **{k: v for k, v in created_org.items() if k != "_id"},
        total_taxistas=0,
        total_vehiculos=0,
        total_clientes=0
    )

@api_router.get("/organizations", response_model=List[OrganizationResponse])
async def get_organizations(
    current_user: dict = Depends(get_current_superadmin),
    activa: Optional[bool] = Query(None, description="Filtrar por estado activo/inactivo")
):
    """Listar todas las organizaciones (solo superadmin) - Optimizado para evitar N+1 queries"""
    query = {}
    if activa is not None:
        query["activa"] = activa
    
    organizations = await db.organizations.find(query).sort("created_at", -1).to_list(1000)
    
    if not organizations:
        return []
    
    # Optimización: obtener todas las estadísticas en 3 queries en lugar de N*3 queries
    org_ids = [str(org["_id"]) for org in organizations]
    
    # Query 1: Contar todos los taxistas por organización
    taxistas_pipeline = [
        {"$match": {"organization_id": {"$in": org_ids}, "role": "taxista"}},
        {"$group": {"_id": "$organization_id", "count": {"$sum": 1}}}
    ]
    taxistas_counts = await db.users.aggregate(taxistas_pipeline).to_list(1000)
    taxistas_map = {c["_id"]: c["count"] for c in taxistas_counts}
    
    # Query 2: Contar todos los vehículos por organización
    vehiculos_pipeline = [
        {"$match": {"organization_id": {"$in": org_ids}}},
        {"$group": {"_id": "$organization_id", "count": {"$sum": 1}}}
    ]
    vehiculos_counts = await db.vehiculos.aggregate(vehiculos_pipeline).to_list(1000)
    vehiculos_map = {c["_id"]: c["count"] for c in vehiculos_counts}
    
    # Query 3: Contar todas las empresas cliente por organización
    clientes_pipeline = [
        {"$match": {"organization_id": {"$in": org_ids}}},
        {"$group": {"_id": "$organization_id", "count": {"$sum": 1}}}
    ]
    clientes_counts = await db.companies.aggregate(clientes_pipeline).to_list(1000)
    clientes_map = {c["_id"]: c["count"] for c in clientes_counts}
    
    # Construir la respuesta usando los mapas precalculados
    result = []
    for org in organizations:
        org_id = str(org["_id"])
        # Asegurar que el campo nombre existe (fix para testing)
        org_data = {k: v for k, v in org.items() if k != "_id"}
        if "nombre" not in org_data or not org_data["nombre"]:
            org_data["nombre"] = f"Organización {org_id[:8]}"
        
        result.append(OrganizationResponse(
            id=org_id,
            **org_data,
            total_taxistas=taxistas_map.get(org_id, 0),
            total_vehiculos=vehiculos_map.get(org_id, 0),
            total_clientes=clientes_map.get(org_id, 0)
        ))
    
    return result

@api_router.get("/organizations/{org_id}", response_model=OrganizationResponse)
async def get_organization(org_id: str, current_user: dict = Depends(get_current_superadmin)):
    """Obtener detalle de una organización (solo superadmin)"""
    org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    if not org:
        raise HTTPException(status_code=404, detail="Organización no encontrada")
    
    # Contar estadísticas
    total_taxistas = await db.users.count_documents({"organization_id": org_id, "role": "taxista"})
    total_vehiculos = await db.vehiculos.count_documents({"organization_id": org_id})
    total_clientes = await db.companies.count_documents({"organization_id": org_id})
    
    return OrganizationResponse(
        id=str(org["_id"]),
        **{k: v for k, v in org.items() if k != "_id"},
        total_taxistas=total_taxistas,
        total_vehiculos=total_vehiculos,
        total_clientes=total_clientes
    )

@api_router.put("/organizations/{org_id}", response_model=OrganizationResponse)
async def update_organization(org_id: str, org_update: OrganizationUpdate, current_user: dict = Depends(get_current_superadmin)):
    """Actualizar organización (solo superadmin)"""
    existing = await db.organizations.find_one({"_id": ObjectId(org_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Organización no encontrada")
    
    update_data = {k: v for k, v in org_update.dict().items() if v is not None}
    
    # Si se actualiza el nombre y no el slug, regenerar slug
    if "nombre" in update_data and "slug" not in update_data:
        update_data["slug"] = generate_slug(update_data["nombre"])
    
    update_data["updated_at"] = datetime.utcnow()
    
    await db.organizations.update_one(
        {"_id": ObjectId(org_id)},
        {"$set": update_data}
    )
    
    updated_org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    
    # Contar estadísticas
    total_taxistas = await db.users.count_documents({"organization_id": org_id, "role": "taxista"})
    total_vehiculos = await db.vehiculos.count_documents({"organization_id": org_id})
    total_clientes = await db.companies.count_documents({"organization_id": org_id})
    
    return OrganizationResponse(
        id=str(updated_org["_id"]),
        **{k: v for k, v in updated_org.items() if k != "_id"},
        total_taxistas=total_taxistas,
        total_vehiculos=total_vehiculos,
        total_clientes=total_clientes
    )

@api_router.delete("/organizations/{org_id}")
async def delete_organization(org_id: str, current_user: dict = Depends(get_current_superadmin)):
    """Eliminar organización y todos sus datos (solo superadmin)"""
    org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    if not org:
        raise HTTPException(status_code=404, detail="Organización no encontrada")
    
    # Eliminar todos los datos relacionados
    deleted_users = await db.users.delete_many({"organization_id": org_id})
    deleted_companies = await db.companies.delete_many({"organization_id": org_id})
    deleted_vehiculos = await db.vehiculos.delete_many({"organization_id": org_id})
    deleted_turnos = await db.turnos.delete_many({"organization_id": org_id})
    deleted_services = await db.services.delete_many({"organization_id": org_id})
    
    # Eliminar la organización
    await db.organizations.delete_one({"_id": ObjectId(org_id)})
    
    return {
        "message": f"Organización '{org['nombre']}' eliminada correctamente",
        "deleted": {
            "users": deleted_users.deleted_count,
            "companies": deleted_companies.deleted_count,
            "vehiculos": deleted_vehiculos.deleted_count,
            "turnos": deleted_turnos.deleted_count,
            "services": deleted_services.deleted_count
        }
    }

# Endpoint para crear admin de organización (superadmin)
@api_router.post("/organizations/{org_id}/admin", response_model=UserResponse)
async def create_organization_admin(
    org_id: str,
    user: UserCreate,
    current_user: dict = Depends(get_current_superadmin)
):
    """Crear administrador para una organización (solo superadmin)"""
    # Verificar que la organización existe
    org = await db.organizations.find_one({"_id": ObjectId(org_id)})
    if not org:
        raise HTTPException(status_code=404, detail="Organización no encontrada")
    
    # Verificar username único
    existing_user = await db.users.find_one({"username": user.username})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    user_dict = user.dict()
    user_dict["password"] = get_password_hash(user_dict["password"])
    user_dict["role"] = "admin"  # Forzar rol admin
    user_dict["organization_id"] = org_id
    user_dict["created_at"] = datetime.utcnow()
    
    result = await db.users.insert_one(user_dict)
    created_user = await db.users.find_one({"_id": result.inserted_id})
    
    return UserResponse(
        id=str(created_user["_id"]),
        username=created_user["username"],
        nombre=created_user["nombre"],
        role=created_user["role"],
        organization_id=created_user.get("organization_id"),
        organization_nombre=org.get("nombre"),
        created_at=created_user["created_at"]
    )

# Endpoint para obtener usuarios sin organización (superadmin)
@api_router.get("/users/unassigned")
async def get_unassigned_users(current_user: dict = Depends(get_current_superadmin)):
    """Obtener usuarios que no tienen organización asignada (solo superadmin)"""
    users = await db.users.find({
        "organization_id": {"$in": [None, ""]},
        "role": {"$ne": "superadmin"}
    }).to_list(1000)
    
    result = []
    for u in users:
        result.append({
            "id": str(u["_id"]),
            "username": u.get("username"),
            "nombre": u.get("nombre"),
            "role": u.get("role"),
            "created_at": u.get("created_at"),
        })
    
    return result

# Endpoint para asignar usuario a organización (superadmin)
@api_router.put("/users/{user_id}/assign-organization/{org_id}")
async def assign_user_to_organization(
    user_id: str,
    org_id: str,
    current_user: dict = Depends(get_current_superadmin)
):
    """Asignar un usuario a una organización (solo superadmin)"""
    # ROBUSTEZ: Validar IDs antes de usar
    user_oid = _get_object_id_or_400(user_id, "user_id")
    org_oid = _get_object_id_or_400(org_id, "org_id")
    
    # Verificar que el usuario existe
    user = await db.users.find_one({"_id": user_oid})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    if user.get("role") == "superadmin":
        raise HTTPException(status_code=400, detail="No se puede asignar organización a superadmin")
    
    # Verificar que la organización existe
    org = await db.organizations.find_one({"_id": org_oid})
    if not org:
        raise HTTPException(status_code=404, detail="Organización no encontrada")
    
    # Actualizar el usuario
    await db.users.update_one(
        {"_id": user_oid},
        {"$set": {"organization_id": org_id}}
    )
    
    # BUG FIX: Migrar datos relacionados usando taxista_id (no user_id)
    # En el modelo de datos, turnos y servicios usan taxista_id para referenciar al usuario
    await db.turnos.update_many(
        {"taxista_id": user_id, "organization_id": {"$in": [None, ""]}},
        {"$set": {"organization_id": org_id}}
    )
    
    await db.services.update_many(
        {"taxista_id": user_id, "organization_id": {"$in": [None, ""]}},
        {"$set": {"organization_id": org_id}}
    )
    
    return {
        "message": f"Usuario '{user.get('nombre')}' asignado a '{org.get('nombre')}'",
        "user_id": user_id,
        "organization_id": org_id,
        "organization_nombre": org.get("nombre")
    }

# Modelo para cambio de contraseña
class PasswordChange(BaseModel):
    new_password: str = Field(..., min_length=6, description="Nueva contraseña (mínimo 6 caracteres)")

# Endpoint para cambiar contraseña de un usuario (superadmin)
@api_router.put("/superadmin/users/{user_id}/change-password")
async def superadmin_change_user_password(
    user_id: str,
    password_data: PasswordChange,
    current_user: dict = Depends(get_current_superadmin)
):
    """Cambiar contraseña de cualquier usuario (solo superadmin)"""
    # Validar ID
    user_oid = _get_object_id_or_400(user_id, "user_id")
    
    # Verificar que el usuario existe
    user = await db.users.find_one({"_id": user_oid})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Encriptar nueva contraseña
    hashed_password = get_password_hash(password_data.new_password)
    
    # Actualizar contraseña
    await db.users.update_one(
        {"_id": user_oid},
        {"$set": {"password": hashed_password, "updated_at": get_spain_now()}}
    )
    
    return {
        "message": f"Contraseña de '{user.get('nombre')}' actualizada correctamente",
        "user_id": user_id,
        "username": user.get("username")
    }

# Endpoint para eliminar un usuario (superadmin)
@api_router.delete("/superadmin/users/{user_id}")
async def superadmin_delete_user(
    user_id: str,
    current_user: dict = Depends(get_current_superadmin)
):
    """Eliminar cualquier usuario excepto superadmin (solo superadmin)"""
    # Validar ID
    user_oid = _get_object_id_or_400(user_id, "user_id")
    
    # Verificar que el usuario existe
    user = await db.users.find_one({"_id": user_oid})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # No permitir eliminar superadmin
    if user.get("role") == "superadmin":
        raise HTTPException(status_code=403, detail="No se puede eliminar un superadmin")
    
    # No permitir eliminarse a sí mismo
    if str(user["_id"]) == str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="No puedes eliminarte a ti mismo")
    
    # Eliminar usuario
    await db.users.delete_one({"_id": user_oid})
    
    return {
        "message": f"Usuario '{user.get('nombre')}' eliminado correctamente",
        "user_id": user_id,
        "username": user.get("username")
    }

# ==========================================
# SUPERADMIN - GESTIÓN DE TAXISTAS
# ==========================================
@api_router.get("/superadmin/admins")
async def superadmin_list_admins(current_user: dict = Depends(get_current_superadmin)):
    """Listar todos los administradores de organizaciones (solo superadmin)"""
    admins = await db.users.find({"role": "admin"}).to_list(1000)
    
    # Get organization names
    org_ids = list(set([a.get("organization_id") for a in admins if a.get("organization_id")]))
    orgs = await db.organizations.find({"_id": {"$in": [ObjectId(oid) for oid in org_ids if oid]}}).to_list(100)
    org_map = {str(o["_id"]): o.get("nombre") for o in orgs}
    
    result = []
    for a in admins:
        result.append({
            "id": str(a["_id"]),
            "username": a.get("username"),
            "nombre": a.get("nombre"),
            "email": a.get("email"),
            "telefono": a.get("telefono"),
            "organization_id": a.get("organization_id"),
            "organization_nombre": org_map.get(a.get("organization_id"), "Sin asignar"),
            "created_at": a.get("created_at")
        })
    return result

@api_router.get("/superadmin/taxistas")
async def superadmin_list_taxistas(current_user: dict = Depends(get_current_superadmin)):
    """Listar todos los taxistas de todas las organizaciones (solo superadmin)"""
    taxistas = await db.users.find({"role": "taxista"}).to_list(1000)
    
    # Get organization names
    org_ids = list(set([t.get("organization_id") for t in taxistas if t.get("organization_id")]))
    orgs = await db.organizations.find({"_id": {"$in": [ObjectId(oid) for oid in org_ids if oid]}}).to_list(100)
    org_map = {str(o["_id"]): o.get("nombre") for o in orgs}
    
    result = []
    for t in taxistas:
        # Compatibilidad: buscar en ambos campos (vehiculo_asignado_* y vehiculo_*)
        # Priorizar vehiculo_asignado_* (nuevo) pero también aceptar vehiculo_* (legacy/admin)
        vehiculo_id = t.get("vehiculo_asignado_id") or t.get("vehiculo_id")
        vehiculo_matricula = t.get("vehiculo_asignado_matricula") or t.get("vehiculo_matricula")
        
        result.append({
            "id": str(t["_id"]),
            "username": t.get("username"),
            "nombre": t.get("nombre"),
            "telefono": t.get("telefono"),
            "email": t.get("email"),
            "licencia": t.get("licencia"),  # Campo de licencia del taxista
            "organization_id": t.get("organization_id"),
            "organization_nombre": org_map.get(t.get("organization_id"), "Sin asignar"),
            "vehiculo_asignado_id": vehiculo_id,
            "vehiculo_asignado_matricula": vehiculo_matricula,
            "activo": t.get("activo", True),
            "created_at": t.get("created_at")
        })
    return result

@api_router.post("/superadmin/taxistas")
async def superadmin_create_taxista(
    taxista: dict,
    current_user: dict = Depends(get_current_superadmin)
):
    """Crear taxista en cualquier organización (solo superadmin)"""
    # Validar campos requeridos
    if not taxista.get("username") or not taxista.get("password") or not taxista.get("nombre"):
        raise HTTPException(status_code=400, detail="Username, password y nombre son obligatorios")
    if not taxista.get("organization_id"):
        raise HTTPException(status_code=400, detail="Debe seleccionar una organización")
    
    # Verificar que el username no existe
    existing = await db.users.find_one({"username": taxista["username"]})
    if existing:
        raise HTTPException(status_code=400, detail="El username ya existe")
    
    # Verificar que la organización existe
    org = await db.organizations.find_one({"_id": ObjectId(taxista["organization_id"])})
    if not org:
        raise HTTPException(status_code=404, detail="Organización no encontrada")
    
    # Incluir TODOS los campos que usa el admin para compatibilidad completa
    user_dict = {
        "username": taxista["username"],
        "password": get_password_hash(taxista["password"]),
        "nombre": taxista["nombre"],
        "telefono": taxista.get("telefono", ""),
        "email": taxista.get("email", ""),
        "licencia": taxista.get("licencia", ""),  # Campo de licencia del taxista
        "role": "taxista",
        "activo": taxista.get("activo", True),
        "organization_id": taxista["organization_id"],
        "created_at": datetime.utcnow()
    }
    
    result = await db.users.insert_one(user_dict)
    return {"id": str(result.inserted_id), "message": "Taxista creado correctamente"}

@api_router.put("/superadmin/taxistas/{taxista_id}")
async def superadmin_update_taxista(
    taxista_id: str,
    taxista: dict,
    current_user: dict = Depends(get_current_superadmin)
):
    """Actualizar taxista (solo superadmin)"""
    existing = await db.users.find_one({"_id": ObjectId(taxista_id), "role": "taxista"})
    if not existing:
        raise HTTPException(status_code=404, detail="Taxista no encontrado")
    
    update_data = {}
    if taxista.get("nombre"):
        update_data["nombre"] = taxista["nombre"]
    if taxista.get("telefono") is not None:
        update_data["telefono"] = taxista["telefono"]
    if taxista.get("email") is not None:
        update_data["email"] = taxista["email"]
    if taxista.get("licencia") is not None:
        update_data["licencia"] = taxista["licencia"]
    if taxista.get("activo") is not None:
        update_data["activo"] = taxista["activo"]
    if taxista.get("organization_id"):
        # Verificar que la organización existe
        org = await db.organizations.find_one({"_id": ObjectId(taxista["organization_id"])})
        if not org:
            raise HTTPException(status_code=404, detail="Organización no encontrada")
        update_data["organization_id"] = taxista["organization_id"]
    if taxista.get("password"):
        update_data["password"] = get_password_hash(taxista["password"])
    
    if update_data:
        await db.users.update_one({"_id": ObjectId(taxista_id)}, {"$set": update_data})
    
    return {"message": "Taxista actualizado correctamente"}

@api_router.delete("/superadmin/taxistas/{taxista_id}")
async def superadmin_delete_taxista(taxista_id: str, current_user: dict = Depends(get_current_superadmin)):
    """Eliminar taxista (solo superadmin)"""
    existing = await db.users.find_one({"_id": ObjectId(taxista_id), "role": "taxista"})
    if not existing:
        raise HTTPException(status_code=404, detail="Taxista no encontrado")
    
    await db.users.delete_one({"_id": ObjectId(taxista_id)})
    return {"message": "Taxista eliminado correctamente"}

@api_router.put("/superadmin/taxistas/{taxista_id}/vehiculo")
async def superadmin_assign_vehiculo_to_taxista(
    taxista_id: str,
    data: dict,
    current_user: dict = Depends(get_current_superadmin)
):
    """Asignar o desasignar vehículo a taxista (solo superadmin)"""
    taxista = await db.users.find_one({"_id": ObjectId(taxista_id), "role": "taxista"})
    if not taxista:
        raise HTTPException(status_code=404, detail="Taxista no encontrado")
    
    vehiculo_id = data.get("vehiculo_id")
    
    # Si hay un vehículo anterior asignado, quitarlo (buscar en ambos campos por compatibilidad)
    old_vehiculo_id = taxista.get("vehiculo_asignado_id") or taxista.get("vehiculo_id")
    if old_vehiculo_id:
        await db.vehiculos.update_one(
            {"_id": ObjectId(old_vehiculo_id)},
            {"$unset": {"taxista_asignado_id": "", "taxista_asignado_nombre": ""}}
        )
    
    if vehiculo_id:
        # Verificar que el vehículo existe y es de la misma organización
        vehiculo = await db.vehiculos.find_one({"_id": ObjectId(vehiculo_id)})
        if not vehiculo:
            raise HTTPException(status_code=404, detail="Vehículo no encontrado")
        if vehiculo.get("organization_id") != taxista.get("organization_id"):
            raise HTTPException(status_code=400, detail="El vehículo debe ser de la misma organización")
        
        # Quitar el vehículo de otro taxista si estaba asignado
        if vehiculo.get("taxista_asignado_id") and vehiculo.get("taxista_asignado_id") != taxista_id:
            await db.users.update_one(
                {"_id": ObjectId(vehiculo["taxista_asignado_id"])},
                {"$unset": {
                    "vehiculo_asignado_id": "", "vehiculo_asignado_matricula": "",
                    "vehiculo_id": "", "vehiculo_matricula": ""  # También limpiar campos legacy
                }}
            )
        
        # Asignar vehículo al taxista (guardar en AMBOS campos para compatibilidad)
        await db.users.update_one(
            {"_id": ObjectId(taxista_id)},
            {"$set": {
                "vehiculo_asignado_id": vehiculo_id, 
                "vehiculo_asignado_matricula": vehiculo.get("matricula"),
                "vehiculo_id": vehiculo_id,  # Campo legacy usado por admin de org
                "vehiculo_matricula": vehiculo.get("matricula")  # Campo legacy
            }}
        )
        
        # Asignar taxista al vehículo
        await db.vehiculos.update_one(
            {"_id": ObjectId(vehiculo_id)},
            {"$set": {"taxista_asignado_id": taxista_id, "taxista_asignado_nombre": taxista.get("nombre")}}
        )
        
        return {"message": f"Vehículo {vehiculo.get('matricula')} asignado a {taxista.get('nombre')}"}
    else:
        # Desasignar vehículo (limpiar AMBOS campos)
        await db.users.update_one(
            {"_id": ObjectId(taxista_id)},
            {"$unset": {
                "vehiculo_asignado_id": "", "vehiculo_asignado_matricula": "",
                "vehiculo_id": "", "vehiculo_matricula": ""  # También limpiar campos legacy
            }}
        )
        return {"message": "Vehículo desasignado"}

# ==========================================
# SUPERADMIN - GESTIÓN DE VEHÍCULOS
# ==========================================
@api_router.get("/superadmin/vehiculos")
async def superadmin_list_vehiculos(current_user: dict = Depends(get_current_superadmin)):
    """Listar todos los vehículos de todas las organizaciones (solo superadmin)"""
    vehiculos = await db.vehiculos.find().to_list(1000)
    
    # Get organization names
    org_ids = list(set([v.get("organization_id") for v in vehiculos if v.get("organization_id")]))
    orgs = await db.organizations.find({"_id": {"$in": [ObjectId(oid) for oid in org_ids if oid]}}).to_list(100)
    org_map = {str(o["_id"]): o.get("nombre") for o in orgs}
    
    result = []
    for v in vehiculos:
        result.append({
            "id": str(v["_id"]),
            "matricula": v.get("matricula"),
            "marca": v.get("marca"),
            "modelo": v.get("modelo"),
            "licencia": v.get("licencia"),
            "plazas": v.get("plazas", 4),
            "km_iniciales": v.get("km_iniciales", 0),
            "fecha_compra": v.get("fecha_compra", ""),
            "activo": v.get("activo", True),
            "organization_id": v.get("organization_id"),
            "organization_nombre": org_map.get(v.get("organization_id"), "Sin asignar"),
            "taxista_asignado_id": v.get("taxista_asignado_id"),
            "taxista_asignado_nombre": v.get("taxista_asignado_nombre")
        })
    return result

@api_router.post("/superadmin/vehiculos")
async def superadmin_create_vehiculo(
    vehiculo: dict,
    current_user: dict = Depends(get_current_superadmin)
):
    """Crear vehiculo en cualquier organizacion (solo superadmin)"""
    if not vehiculo.get("matricula"):
        raise HTTPException(status_code=400, detail="La matricula es obligatoria")
    if not vehiculo.get("organization_id"):
        raise HTTPException(status_code=400, detail="Debe seleccionar una organizacion")
    
    # Verificar que la organizacion existe
    org = await db.organizations.find_one({"_id": ObjectId(vehiculo["organization_id"])})
    if not org:
        raise HTTPException(status_code=404, detail="Organizacion no encontrada")
    
    # Verificar que la matricula no exista en esta organizacion
    matricula_upper = vehiculo["matricula"].upper()
    existing = await db.vehiculos.find_one({
        "matricula": matricula_upper,
        "organization_id": vehiculo["organization_id"]
    })
    if existing:
        raise HTTPException(status_code=400, detail="La matricula ya existe en esta organizacion")
    
    # Incluir TODOS los campos que usa el admin para compatibilidad completa
    vehiculo_dict = {
        "matricula": matricula_upper,
        "marca": vehiculo.get("marca", ""),
        "modelo": vehiculo.get("modelo", ""),
        "licencia": vehiculo.get("licencia", ""),
        "plazas": vehiculo.get("plazas", 4),  # Default 4 plazas
        "km_iniciales": vehiculo.get("km_iniciales", 0),
        "fecha_compra": vehiculo.get("fecha_compra", ""),
        "activo": vehiculo.get("activo", True),
        "organization_id": vehiculo["organization_id"],
        "created_at": datetime.utcnow()
    }
    
    result = await db.vehiculos.insert_one(vehiculo_dict)
    return {"id": str(result.inserted_id), "message": "Vehiculo creado correctamente"}

@api_router.put("/superadmin/vehiculos/{vehiculo_id}")
async def superadmin_update_vehiculo(
    vehiculo_id: str,
    vehiculo: dict,
    current_user: dict = Depends(get_current_superadmin)
):
    """Actualizar vehículo (solo superadmin)"""
    existing = await db.vehiculos.find_one({"_id": ObjectId(vehiculo_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    
    update_data = {}
    if vehiculo.get("matricula"):
        update_data["matricula"] = vehiculo["matricula"].upper()
    if vehiculo.get("marca") is not None:
        update_data["marca"] = vehiculo["marca"]
    if vehiculo.get("modelo") is not None:
        update_data["modelo"] = vehiculo["modelo"]
    if vehiculo.get("licencia") is not None:
        update_data["licencia"] = vehiculo["licencia"]
    if vehiculo.get("plazas") is not None:
        update_data["plazas"] = vehiculo["plazas"]
    if vehiculo.get("km_iniciales") is not None:
        update_data["km_iniciales"] = vehiculo["km_iniciales"]
    if vehiculo.get("fecha_compra") is not None:
        update_data["fecha_compra"] = vehiculo["fecha_compra"]
    if vehiculo.get("activo") is not None:
        update_data["activo"] = vehiculo["activo"]
    if vehiculo.get("organization_id"):
        org = await db.organizations.find_one({"_id": ObjectId(vehiculo["organization_id"])})
        if not org:
            raise HTTPException(status_code=404, detail="Organización no encontrada")
        update_data["organization_id"] = vehiculo["organization_id"]
    
    if update_data:
        await db.vehiculos.update_one({"_id": ObjectId(vehiculo_id)}, {"$set": update_data})
    
    return {"message": "Vehículo actualizado correctamente"}

@api_router.delete("/superadmin/vehiculos/{vehiculo_id}")
async def superadmin_delete_vehiculo(vehiculo_id: str, current_user: dict = Depends(get_current_superadmin)):
    """Eliminar vehículo (solo superadmin)"""
    existing = await db.vehiculos.find_one({"_id": ObjectId(vehiculo_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    
    await db.vehiculos.delete_one({"_id": ObjectId(vehiculo_id)})
    return {"message": "Vehículo eliminado correctamente"}

# ==========================================
# USER ENDPOINTS (Multi-tenant - admin only)
# ==========================================
@api_router.post("/users", response_model=UserResponse)
async def create_user(user: UserCreate, current_user: dict = Depends(get_current_admin)):
    """
    Crear taxista - se asigna automáticamente a la organización del admin.
    SEGURIDAD:
    - Admin solo puede crear taxistas (no admin ni superadmin)
    - Valida vehiculo_id pertenece a la misma organización
    - vehiculo_matricula se obtiene de BD (ignora cliente)
    """
    # SEGURIDAD: Admin no puede crear admins ni superadmins
    if not is_superadmin(current_user) and user.role in ["admin", "superadmin"]:
        raise HTTPException(
            status_code=403, 
            detail="No tienes permiso para crear usuarios con rol admin o superadmin"
        )
    
    # SEGURIDAD: Nadie puede crear superadmin desde este endpoint
    if user.role == "superadmin":
        raise HTTPException(
            status_code=403, 
            detail="No se puede crear usuarios con rol superadmin desde este endpoint"
        )
    
    # Check if user exists
    existing_user = await db.users.find_one({"username": user.username})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Username already exists"
        )
    
    # Determinar organization_id
    org_id = None
    if not is_superadmin(current_user):
        org_id = get_user_organization_id(current_user)
    elif user.organization_id:
        org_id = user.organization_id
    
    # INTEGRIDAD: Validar vehiculo_id y obtener matricula desde BD
    vehiculo_validated = None
    if user.vehiculo_id:
        try:
            vehiculo_query = {"_id": ObjectId(user.vehiculo_id)}
            if org_id:
                vehiculo_query["organization_id"] = org_id
            vehiculo_validated = await db.vehiculos.find_one(vehiculo_query)
            if not vehiculo_validated:
                raise HTTPException(
                    status_code=400, 
                    detail="El vehículo especificado no existe o no pertenece a esta organización"
                )
        except Exception:
            raise HTTPException(status_code=400, detail="vehiculo_id inválido")
    
    user_dict = user.dict()
    user_dict["password"] = get_password_hash(user_dict["password"])
    user_dict["created_at"] = datetime.utcnow()
    user_dict["organization_id"] = org_id
    
    # INTEGRIDAD: Usar matrícula desde BD, ignorar lo que venga del cliente
    if vehiculo_validated:
        user_dict["vehiculo_id"] = str(vehiculo_validated["_id"])
        user_dict["vehiculo_matricula"] = vehiculo_validated.get("matricula", "")
    else:
        # Si no hay vehiculo válido, limpiar para evitar inconsistencias
        user_dict["vehiculo_id"] = None
        user_dict["vehiculo_matricula"] = None
    
    result = await db.users.insert_one(user_dict)
    created_user = await db.users.find_one({"_id": result.inserted_id})
    
    # Obtener nombre de organización
    org_nombre = None
    if created_user.get("organization_id"):
        org = await db.organizations.find_one({"_id": ObjectId(created_user["organization_id"])})
        if org:
            org_nombre = org.get("nombre")
    
    return UserResponse(
        id=str(created_user["_id"]),
        username=created_user["username"],
        nombre=created_user["nombre"],
        role=created_user["role"],
        licencia=created_user.get("licencia"),
        vehiculo_id=created_user.get("vehiculo_id"),
        vehiculo_matricula=created_user.get("vehiculo_matricula"),
        organization_id=created_user.get("organization_id"),
        organization_nombre=org_nombre,
        created_at=created_user["created_at"]
    )

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(get_current_admin)):
    """Listar taxistas - filtrado por organización del admin"""
    # Multi-tenant filter
    org_filter = await get_org_filter(current_user)
    query = {"role": "taxista", **org_filter}
    
    users = await db.users.find(
        query,
        {"password": 0}
    ).to_list(1000)
    
    # Obtener nombres de organizaciones
    org_names = {}
    org_ids = set(u.get("organization_id") for u in users if u.get("organization_id"))
    for org_id in org_ids:
        org = await db.organizations.find_one({"_id": ObjectId(org_id)})
        if org:
            org_names[org_id] = org.get("nombre")
    
    return [
        UserResponse(
            id=str(user["_id"]),
            username=user["username"],
            nombre=user["nombre"],
            role=user["role"],
            licencia=user.get("licencia"),
            vehiculo_id=user.get("vehiculo_id"),
            vehiculo_matricula=user.get("vehiculo_matricula"),
            organization_id=user.get("organization_id"),
            organization_nombre=org_names.get(user.get("organization_id")),
            created_at=user["created_at"]
        )
        for user in users
    ]

class UserUpdate(BaseModel):
    username: str
    nombre: str
    role: Optional[str] = None  # SEGURIDAD: No se permite cambiar rol desde aquí
    licencia: Optional[str] = None
    vehiculo_id: Optional[str] = None
    vehiculo_matricula: Optional[str] = None
    password: Optional[str] = None

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(user_id: str, user: UserUpdate, current_user: dict = Depends(get_current_admin)):
    # SEGURIDAD: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    
    # Verificar que el usuario existe y pertenece a la organización
    existing_user = await db.users.find_one({"_id": ObjectId(user_id), **org_filter})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # SEGURIDAD P0: Bloquear cambio de roles para admins
    # Solo superadmin puede cambiar roles (y lo hace desde otro endpoint)
    user_dict = user.dict(exclude={'password', 'role'}, exclude_none=True)  # Excluir role siempre
    
    # INTEGRIDAD: Validar que vehiculo_id pertenece a la misma organización
    if user.vehiculo_id:
        org_id = existing_user.get("organization_id")
        vehiculo_query = {"_id": ObjectId(user.vehiculo_id)}
        if org_id:
            vehiculo_query["organization_id"] = org_id
        vehiculo = await db.vehiculos.find_one(vehiculo_query)
        if not vehiculo:
            raise HTTPException(
                status_code=400, 
                detail="El vehículo especificado no existe o no pertenece a esta organización"
            )
        # Actualizar también la matrícula para mantener consistencia
        user_dict["vehiculo_matricula"] = vehiculo.get("matricula")
    
    # Si se proporciona una nueva contraseña, hashearla
    if user.password:
        user_dict["password"] = get_password_hash(user.password)
    
    result = await db.users.update_one(
        {"_id": ObjectId(user_id), **org_filter},  # Doble check
        {"$set": user_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    updated_user = await db.users.find_one({"_id": ObjectId(user_id), **org_filter})
    return UserResponse(
        id=str(updated_user["_id"]),
        username=updated_user["username"],
        nombre=updated_user["nombre"],
        role=updated_user["role"],
        licencia=updated_user.get("licencia"),
        vehiculo_id=updated_user.get("vehiculo_id"),
        vehiculo_matricula=updated_user.get("vehiculo_matricula"),
        created_at=updated_user["created_at"]
    )

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_admin)):
    # SEGURIDAD: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    
    result = await db.users.delete_one({"_id": ObjectId(user_id), **org_filter})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted successfully"}

# ==========================================
# COMPANY/CLIENT ENDPOINTS (Multi-tenant)
# ==========================================
@api_router.post("/companies", response_model=CompanyResponse)
async def create_company(company: CompanyCreate, current_user: dict = Depends(get_current_admin)):
    """Crear cliente/empresa - se asigna automaticamente a la organizacion del admin"""
    
    # SEGURIDAD: Superadmin no puede crear empresas (evita datos sin tenant)
    if is_superadmin(current_user):
        raise HTTPException(
            status_code=403,
            detail="Superadmin no puede crear empresas. Use una cuenta de admin de organizacion."
        )
    
    company_dict = company.dict()
    
    # Multi-tenant: Asignar organization_id
    org_id = get_user_organization_id(current_user)
    company_dict["organization_id"] = org_id
    
    # Validar numero_cliente unico dentro de la organizacion
    if company_dict.get("numero_cliente"):
        query = {"numero_cliente": company_dict["numero_cliente"], "organization_id": org_id}
        existing = await db.companies.find_one(query)
        if existing:
            raise HTTPException(status_code=400, detail="El numero de cliente ya existe en tu organizacion")
    
    company_dict["created_at"] = datetime.utcnow()
    
    result = await db.companies.insert_one(company_dict)
    created_company = await db.companies.find_one({"_id": result.inserted_id})
    
    return CompanyResponse(
        id=str(created_company["_id"]),
        **{k: v for k, v in created_company.items() if k != "_id"}
    )

@api_router.get("/companies", response_model=List[CompanyResponse])
async def get_companies(current_user: dict = Depends(get_current_user)):
    """Listar clientes/empresas - filtrado por organización"""
    # Multi-tenant filter
    org_filter = await get_org_filter(current_user)
    
    companies = await db.companies.find(org_filter).to_list(1000)
    return [
        CompanyResponse(
            id=str(company["_id"]),
            **{k: v for k, v in company.items() if k != "_id"}
        )
        for company in companies
    ]

@api_router.put("/companies/{company_id}", response_model=CompanyResponse)
async def update_company(company_id: str, company: CompanyCreate, current_user: dict = Depends(get_current_admin)):
    """Actualizar cliente/empresa - solo de la propia organización"""
    company_dict = company.dict()
    
    # Multi-tenant: Verificar que la empresa pertenece a la organización del usuario
    org_filter = await get_org_filter(current_user)
    existing_company = await db.companies.find_one({"_id": ObjectId(company_id), **org_filter})
    if not existing_company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Validar numero_cliente único dentro de la organización (excepto el actual)
    if company_dict.get("numero_cliente"):
        query = {
            "numero_cliente": company_dict["numero_cliente"],
            "_id": {"$ne": ObjectId(company_id)},
            **org_filter
        }
        existing = await db.companies.find_one(query)
        if existing:
            raise HTTPException(status_code=400, detail="Número de cliente ya existe")
    
    result = await db.companies.update_one(
        {"_id": ObjectId(company_id), **org_filter},
        {"$set": company_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    
    updated_company = await db.companies.find_one({"_id": ObjectId(company_id), **org_filter})
    return CompanyResponse(
        id=str(updated_company["_id"]),
        **{k: v for k, v in updated_company.items() if k != "_id"}
    )

@api_router.delete("/companies/{company_id}")
async def delete_company(company_id: str, current_user: dict = Depends(get_current_admin)):
    """Eliminar cliente/empresa - solo de la propia organización"""
    # Multi-tenant: Verificar que la empresa pertenece a la organización
    org_filter = await get_org_filter(current_user)
    result = await db.companies.delete_one({"_id": ObjectId(company_id), **org_filter})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    return {"message": "Company deleted successfully"}

# ==========================================
# VEHICULO ENDPOINTS (Multi-tenant)
# ==========================================
@api_router.post("/vehiculos", response_model=VehiculoResponse)
async def create_vehiculo(vehiculo: VehiculoCreate, current_user: dict = Depends(get_current_admin)):
    """Crear vehiculo - se asigna a la organizacion del admin"""
    
    # SEGURIDAD: Superadmin no puede crear vehiculos (evita datos sin tenant)
    if is_superadmin(current_user):
        raise HTTPException(
            status_code=403,
            detail="Superadmin no puede crear vehiculos. Use una cuenta de admin de organizacion."
        )
    
    # Multi-tenant: Asignar organization_id
    org_id = get_user_organization_id(current_user)
    
    # Verificar que la matricula no exista dentro de la organizacion
    query = {"matricula": vehiculo.matricula, "organization_id": org_id}
    existing = await db.vehiculos.find_one(query)
    if existing:
        raise HTTPException(status_code=400, detail="La matricula ya existe en tu organizacion")
    
    vehiculo_dict = vehiculo.dict()
    vehiculo_dict["organization_id"] = org_id
    vehiculo_dict["created_at"] = datetime.utcnow()
    
    result = await db.vehiculos.insert_one(vehiculo_dict)
    created_vehiculo = await db.vehiculos.find_one({"_id": result.inserted_id})
    
    return VehiculoResponse(
        id=str(created_vehiculo["_id"]),
        **{k: v for k, v in created_vehiculo.items() if k != "_id"}
    )

@api_router.get("/vehiculos", response_model=List[VehiculoResponse])
async def get_vehiculos(current_user: dict = Depends(get_current_user)):
    """Listar vehículos - filtrado por organización"""
    org_filter = await get_org_filter(current_user)
    vehiculos = await db.vehiculos.find(org_filter).to_list(1000)
    return [
        VehiculoResponse(
            id=str(vehiculo["_id"]),
            **{k: v for k, v in vehiculo.items() if k != "_id"}
        )
        for vehiculo in vehiculos
    ]

@api_router.put("/vehiculos/{vehiculo_id}", response_model=VehiculoResponse)
async def update_vehiculo(vehiculo_id: str, vehiculo: VehiculoCreate, current_user: dict = Depends(get_current_admin)):
    """Actualizar vehículo - solo de la propia organización"""
    org_filter = await get_org_filter(current_user)
    
    # Verificar que el vehículo pertenece a la organización
    existing_vehiculo = await db.vehiculos.find_one({"_id": ObjectId(vehiculo_id), **org_filter})
    if not existing_vehiculo:
        raise HTTPException(status_code=404, detail="Vehículo not found")
    
    # Verificar que la matrícula no esté en uso por otro vehículo de la organización
    existing = await db.vehiculos.find_one({
        "matricula": vehiculo.matricula,
        "_id": {"$ne": ObjectId(vehiculo_id)},
        **org_filter
    })
    if existing:
        raise HTTPException(status_code=400, detail="La matrícula ya existe")
    
    vehiculo_dict = vehiculo.dict()
    result = await db.vehiculos.update_one(
        {"_id": ObjectId(vehiculo_id), **org_filter},
        {"$set": vehiculo_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vehículo not found")
    
    updated_vehiculo = await db.vehiculos.find_one({"_id": ObjectId(vehiculo_id), **org_filter})
    return VehiculoResponse(
        id=str(updated_vehiculo["_id"]),
        **{k: v for k, v in updated_vehiculo.items() if k != "_id"}
    )

@api_router.delete("/vehiculos/{vehiculo_id}")
async def delete_vehiculo(vehiculo_id: str, current_user: dict = Depends(get_current_admin)):
    """Eliminar vehículo - solo de la propia organización"""
    org_filter = await get_org_filter(current_user)
    result = await db.vehiculos.delete_one({"_id": ObjectId(vehiculo_id), **org_filter})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vehículo not found")
    return {"message": "Vehículo deleted successfully"}

# ==========================================
# TURNO ENDPOINTS (Multi-tenant)
# ==========================================
@api_router.post("/turnos", response_model=TurnoResponse)
async def create_turno(turno: TurnoCreate, current_user: dict = Depends(get_current_user)):
    """Crear turno - se asigna a la organización del usuario"""
    
    # SEGURIDAD P1: Superadmin no puede crear turnos (evita datos sin tenant)
    if is_superadmin(current_user):
        raise HTTPException(
            status_code=403,
            detail="Superadmin no puede crear turnos. Use una cuenta de taxista."
        )
    
    # Validar que no tenga un turno abierto
    existing_turno = await db.turnos.find_one({
        "taxista_id": str(current_user["_id"]),
        "cerrado": False
    })
    if existing_turno:
        raise HTTPException(status_code=400, detail="Ya tienes un turno abierto. Debes finalizarlo antes de abrir uno nuevo.")
    
    org_id = get_user_organization_id(current_user)
    
    # INTEGRIDAD: Validar vehiculo_id y obtener matrícula desde BD (no confiar en cliente)
    vehiculo_validated = None
    if turno.vehiculo_id:
        vehiculo_query = {"_id": ObjectId(turno.vehiculo_id)}
        if org_id:
            vehiculo_query["organization_id"] = org_id
        vehiculo_validated = await db.vehiculos.find_one(vehiculo_query)
        if not vehiculo_validated:
            raise HTTPException(
                status_code=400, 
                detail="El vehículo especificado no existe o no pertenece a esta organización"
            )
    
    turno_dict = turno.dict()
    
    # Override taxista info with current user
    turno_dict["taxista_id"] = str(current_user["_id"])
    turno_dict["taxista_nombre"] = current_user["nombre"]
    turno_dict["created_at"] = datetime.utcnow()
    turno_dict["cerrado"] = False
    
    # (C) HORA DEL SERVIDOR: Usar hora del servidor EN ESPAÑA, ignorar hora_inicio del cliente
    server_now = get_spain_now()
    turno_dict["hora_inicio"] = server_now.strftime("%H:%M")
    
    # Calcular inicio_dt_utc para ordenacion y filtros correctos
    inicio_dt_utc = parse_spanish_date_to_utc(turno_dict["fecha_inicio"], turno_dict["hora_inicio"])
    if inicio_dt_utc:
        turno_dict["inicio_dt_utc"] = inicio_dt_utc
    
    # INTEGRIDAD: Usar matrícula desde BD, ignorar lo que venga del cliente
    if vehiculo_validated:
        turno_dict["vehiculo_id"] = str(vehiculo_validated["_id"])
        turno_dict["vehiculo_matricula"] = vehiculo_validated.get("matricula", "")
    elif "vehiculo_matricula" in turno_dict:
        # Si no hay vehiculo_id válido, limpiar matricula para evitar inconsistencias
        turno_dict["vehiculo_matricula"] = ""
    
    # Inicializar combustible como vacío
    turno_dict["combustible"] = None
    
    # Multi-tenant: Asignar organization_id del usuario
    turno_dict["organization_id"] = org_id
    
    result = await db.turnos.insert_one(turno_dict)
    created_turno = await db.turnos.find_one({"_id": result.inserted_id})
    
    return TurnoResponse(
        id=str(created_turno["_id"]),
        **{k: v for k, v in created_turno.items() if k != "_id"}
    )

# HELPER FUNCTION: Batch fetch servicios para turnos (optimiza N+1 queries)
async def get_turnos_with_servicios(turnos: list, org_filter: dict = None, include_servicios_detail: bool = False) -> list:
    """
    Optimización: Trae todos los servicios de múltiples turnos en 1 query
    en vez de hacer 1 query por cada turno (N+1 problem)
    
    Args:
        turnos: Lista de turnos
        org_filter: Filtro de organización (vacío {} para superadmin, con organization_id para admin/taxista)
        include_servicios_detail: Si True, incluye la lista completa de servicios en cada turno
    
    SEGURIDAD: org_filter evita que servicios de otras organizaciones se mezclen
    aunque tengan turno_id apuntando a turnos de este tenant (contaminación de datos)
    """
    if not turnos:
        return []
    
    # Si no se proporciona org_filter, usar filtro vacío (comportamiento legacy, pero no recomendado)
    if org_filter is None:
        org_filter = {}
    
    # Batch query - traer servicios con filtro de organización para evitar contaminación
    turno_ids = [str(t["_id"]) for t in turnos]
    services_query = {"turno_id": {"$in": turno_ids}, **org_filter}
    all_servicios = await db.services.find(services_query).to_list(MAX_BATCH_SERVICES)
    
    # Guard defensivo: si alcanzamos el límite, devolvemos error controlado
    if len(all_servicios) >= MAX_BATCH_SERVICES:
        raise HTTPException(
            status_code=413,
            detail=f"Demasiados servicios para esta operación batch. Use filtros más específicos o paginación. Límite={MAX_BATCH_SERVICES}"
        )
    
    # Agrupar servicios por turno_id en memoria
    servicios_by_turno = {}
    for servicio in all_servicios:
        turno_id = servicio["turno_id"]
        if turno_id not in servicios_by_turno:
            servicios_by_turno[turno_id] = []
        servicios_by_turno[turno_id].append(servicio)
    
    # Calcular totales para cada turno
    result = []
    for turno in turnos:
        turno_id = str(turno["_id"])
        servicios = servicios_by_turno.get(turno_id, [])
        
        total_clientes = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "empresa")
        total_particulares = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "particular")
        total_km = sum(s.get("kilometros") or 0 for s in servicios)
        
        turno_data = {
            **turno,
            "turno_id": turno_id,
            "total_clientes": total_clientes,
            "total_particulares": total_particulares,
            "total_km": total_km,
            "cantidad_servicios": len(servicios)
        }
        
        # Si se solicita, incluir el detalle completo de servicios
        if include_servicios_detail:
            turno_data["servicios"] = servicios
        
        result.append(turno_data)
    
    return result

@api_router.get("/turnos", response_model=List[TurnoResponse])
async def get_turnos(
    current_user: dict = Depends(get_current_user),
    taxista_id: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    cerrado: Optional[bool] = Query(None),
    liquidado: Optional[bool] = Query(None),
    repostado: Optional[bool] = Query(None, description="Filtrar por turnos con repostaje (combustible)"),
    vehiculo_id: Optional[str] = Query(None, description="Filtrar por vehículo de repostaje"),
    limit: int = Query(500, le=1000, description="Límite de resultados")
):
    """Listar turnos - filtrado por organización"""
    # Multi-tenant filter
    org_filter = await get_org_filter(current_user)
    query = {**org_filter}
    
    # Si no es admin/superadmin, solo sus propios turnos
    if current_user.get("role") not in ["admin", "superadmin"]:
        query["taxista_id"] = str(current_user["_id"])
    elif taxista_id:
        # SEGURIDAD: Validar que taxista_id pertenece al scope antes de filtrar
        await _get_taxista_or_400(taxista_id, org_filter, db)
        query["taxista_id"] = taxista_id
    
    # Filtro por fechas
    if fecha_inicio:
        query["fecha_inicio"] = {"$gte": fecha_inicio}
    if fecha_fin:
        if "fecha_inicio" in query:
            query["fecha_inicio"]["$lte"] = fecha_fin
        else:
            query["fecha_inicio"] = {"$lte": fecha_fin}
    
    # Filtros por estado
    if cerrado is not None:
        query["cerrado"] = cerrado
    if liquidado is not None:
        query["liquidado"] = liquidado
    
    # (F) Filtro por combustible/repostaje
    if repostado is not None:
        query["combustible.repostado"] = repostado
    if vehiculo_id:
        query["combustible.vehiculo_id"] = vehiculo_id
    
    # Validar y ajustar límite
    if limit <= 0:
        limit = 500  # Default
    elif limit > 1000:
        limit = 1000  # Maximum
    
    turnos = await db.turnos.find(query).sort("created_at", -1).limit(limit).to_list(limit)
    
    # OPTIMIZACIÓN: Batch query - traer todos los servicios de una vez
    if turnos:
        turno_ids = [str(t["_id"]) for t in turnos]
        all_servicios = await db.services.find(
            {"turno_id": {"$in": turno_ids}, **org_filter}  # Con org_filter
        ).to_list(MAX_BATCH_SERVICES)
        
        # Guard defensivo: si alcanzamos el límite, devolvemos error controlado
        if len(all_servicios) >= MAX_BATCH_SERVICES:
            raise HTTPException(
                status_code=413,
                detail=f"Demasiados servicios para esta operación batch. Use filtros más específicos o paginación. Límite={MAX_BATCH_SERVICES}"
            )
        
        # Agrupar servicios por turno_id en memoria
        servicios_by_turno = {}
        for servicio in all_servicios:
            turno_id = servicio["turno_id"]
            if turno_id not in servicios_by_turno:
                servicios_by_turno[turno_id] = []
            servicios_by_turno[turno_id].append(servicio)
    else:
        servicios_by_turno = {}
    
    # Calcular totales para cada turno
    result = []
    for turno in turnos:
        turno_id = str(turno["_id"])
        
        # Obtener servicios del turno desde el diccionario
        servicios = servicios_by_turno.get(turno_id, [])
        
        total_clientes = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "empresa")
        total_particulares = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "particular")
        
        # Calcular km del turno: usar km_fin si existe, sino usar suma de km de servicios
        if turno.get("km_fin") is not None:
            total_km = turno["km_fin"] - turno["km_inicio"]
        else:
            total_km = sum(s.get("kilometros") or 0 for s in servicios)
        
        result.append(TurnoResponse(
            id=turno_id,
            **{k: v for k, v in turno.items() if k != "_id"},
            total_importe_clientes=total_clientes,
            total_importe_particulares=total_particulares,
            total_kilometros=total_km,
            cantidad_servicios=len(servicios)
        ))
    
    return result

@api_router.get("/turnos/activo")
async def get_turno_activo(current_user: dict = Depends(get_current_user)):
    turno = await db.turnos.find_one({
        "taxista_id": str(current_user["_id"]),
        "cerrado": False
    })
    
    if not turno:
        return None
    
    turno_id = str(turno["_id"])
    servicios = await db.services.find({"turno_id": turno_id}).to_list(1000)
    
    total_clientes = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "empresa")
    total_particulares = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "particular")
    
    # Calcular km del turno: usar km_fin si existe, sino usar suma de km de servicios
    if turno.get("km_fin") is not None:
        total_km = turno["km_fin"] - turno["km_inicio"]
    else:
        total_km = sum(s.get("kilometros") or 0 for s in servicios)
    
    return TurnoResponse(
        id=turno_id,
        **{k: v for k, v in turno.items() if k != "_id"},
        total_importe_clientes=total_clientes,
        total_importe_particulares=total_particulares,
        total_kilometros=total_km,
        cantidad_servicios=len(servicios)
    )

@api_router.put("/turnos/{turno_id}/finalizar", response_model=TurnoResponse)
async def finalizar_turno(turno_id: str, turno_update: TurnoFinalizarUpdate, current_user: dict = Depends(get_current_user)):
    # SEGURIDAD: Scope por organización
    org_filter = await get_org_filter(current_user)
    
    try:
        oid = ObjectId(turno_id)
    except Exception:
        raise HTTPException(status_code=400, detail="turno_id inválido")
    
    existing_turno = await db.turnos.find_one({"_id": oid, **org_filter})
    if not existing_turno:
        raise HTTPException(status_code=404, detail="Turno not found")
    
    # Solo el taxista dueño o admin/superadmin pueden finalizar
    if current_user.get("role") not in ("admin", "superadmin") and existing_turno["taxista_id"] != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="No autorizado")
    
    update_dict = turno_update.dict()
    
    # (C) HORA DEL SERVIDOR: Usar hora del servidor EN ESPAÑA, ignorar hora_fin del cliente
    server_now = get_spain_now()
    update_dict["hora_fin"] = server_now.strftime("%H:%M")
    
    # Calcular fin_dt_utc para ordenacion y filtros correctos
    fin_dt_utc = parse_spanish_date_to_utc(update_dict.get("fecha_fin"), update_dict["hora_fin"])
    if fin_dt_utc:
        update_dict["fin_dt_utc"] = fin_dt_utc
    
    await db.turnos.update_one(
        {"_id": oid, **org_filter},
        {"$set": update_dict}
    )
    
    updated_turno = await db.turnos.find_one({"_id": oid, **org_filter})
    
    # Calcular totales (scoped)
    servicios = await db.services.find({"turno_id": turno_id, **org_filter}).to_list(1000)
    total_clientes = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "empresa")
    total_particulares = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "particular")
    total_km = sum(s.get("kilometros") or 0 for s in servicios)
    
    return TurnoResponse(
        id=turno_id,
        **{k: v for k, v in updated_turno.items() if k != "_id"},
        total_importe_clientes=total_clientes,
        total_importe_particulares=total_particulares,
        total_kilometros=total_km,
        cantidad_servicios=len(servicios)
    )

@api_router.put("/turnos/{turno_id}", response_model=TurnoResponse)
async def update_turno(turno_id: str, turno_update: TurnoUpdate, current_user: dict = Depends(get_current_admin)):
    """Actualizar turno (admin/superadmin). Permite editar cualquier campo del turno."""
    # SEGURIDAD: Scope por organización
    org_filter = await get_org_filter(current_user)
    
    try:
        oid = ObjectId(turno_id)
    except Exception:
        raise HTTPException(status_code=400, detail="turno_id inválido")
    
    existing_turno = await db.turnos.find_one({"_id": oid, **org_filter})
    if not existing_turno:
        raise HTTPException(status_code=404, detail="Turno not found")
    
    update_dict = turno_update.dict(exclude_none=True)
    if update_dict:
        await db.turnos.update_one(
            {"_id": oid, **org_filter},
            {"$set": update_dict}
        )
    
    updated_turno = await db.turnos.find_one({"_id": oid, **org_filter})
    
    # Calcular totales (scoped)
    servicios = await db.services.find({"turno_id": turno_id, **org_filter}).to_list(1000)
    total_clientes = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "empresa")
    total_particulares = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "particular")
    total_km = sum(s.get("kilometros") or 0 for s in servicios)
    
    return TurnoResponse(
        id=turno_id,
        **{k: v for k, v in updated_turno.items() if k != "_id"},
        total_importe_clientes=total_clientes,
        total_importe_particulares=total_particulares,
        total_kilometros=total_km,
        cantidad_servicios=len(servicios)
    )

@api_router.delete("/turnos/{turno_id}")
async def delete_turno(turno_id: str, current_user: dict = Depends(get_current_admin)):
    """
    Eliminar un turno (admin/superadmin) y sus servicios asociados.
    SEGURIDAD: Scoped por organización.
    """
    # SEGURIDAD: Scope por organización
    org_filter = await get_org_filter(current_user)
    
    try:
        oid = ObjectId(turno_id)
    except Exception:
        raise HTTPException(status_code=400, detail="turno_id inválido")
    
    # Verificar que el turno existe dentro del scope
    turno = await db.turnos.find_one({"_id": oid, **org_filter})
    if not turno:
        raise HTTPException(status_code=404, detail="Turno no encontrado")
    
    # Eliminar todos los servicios asociados al turno (scoped)
    servicios_result = await db.services.delete_many({"turno_id": turno_id, **org_filter})
    
    # Eliminar el turno (scoped)
    await db.turnos.delete_one({"_id": oid, **org_filter})
    
    return {
        "message": "Turno eliminado correctamente",
        "turno_id": turno_id,
        "servicios_eliminados": servicios_result.deleted_count
    }

# (F) COMBUSTIBLE: Endpoint para registrar/editar combustible en turno
@api_router.put("/turnos/{turno_id}/combustible")
async def update_turno_combustible(
    turno_id: str,
    combustible_update: CombustibleUpdate,
    current_user: dict = Depends(get_current_user)
):
    """
    Registrar o editar combustible en un turno.
    - Solo el taxista dueño puede modificar (mientras el turno esté activo)
    - Admin solo puede leer
    - Una vez finalizado el turno, el combustible queda bloqueado
    """
    org_filter = await get_org_filter(current_user)
    org_id = get_user_organization_id(current_user)
    
    try:
        oid = ObjectId(turno_id)
    except Exception:
        raise HTTPException(status_code=400, detail="turno_id inválido")
    
    existing_turno = await db.turnos.find_one({"_id": oid, **org_filter})
    if not existing_turno:
        raise HTTPException(status_code=404, detail="Turno no encontrado")
    
    # Solo el taxista dueño puede modificar combustible
    is_owner = existing_turno["taxista_id"] == str(current_user["_id"])
    if not is_owner:
        raise HTTPException(
            status_code=403,
            detail="Solo el taxista dueño del turno puede registrar combustible"
        )
    
    # Bloqueado si el turno está cerrado
    if existing_turno.get("cerrado", False):
        raise HTTPException(
            status_code=400,
            detail="No se puede modificar combustible en un turno finalizado"
        )
    
    # Validaciones si repostado=True
    if combustible_update.repostado:
        if combustible_update.litros is None or combustible_update.litros <= 0:
            raise HTTPException(
                status_code=400,
                detail="Si repostado=True, litros debe ser > 0"
            )
        if not combustible_update.vehiculo_id:
            raise HTTPException(
                status_code=400,
                detail="Si repostado=True, vehiculo_id es obligatorio"
            )
        if combustible_update.km_vehiculo is None or combustible_update.km_vehiculo < 0:
            raise HTTPException(
                status_code=400,
                detail="Si repostado=True, km_vehiculo es obligatorio y debe ser >= 0"
            )
        
        # Validar que vehiculo_id pertenece al tenant
        try:
            vehiculo_query = {"_id": ObjectId(combustible_update.vehiculo_id)}
            if org_id:
                vehiculo_query["organization_id"] = org_id
            vehiculo = await db.vehiculos.find_one(vehiculo_query)
            if not vehiculo:
                raise HTTPException(
                    status_code=400,
                    detail="vehiculo_id no existe o no pertenece a esta organización"
                )
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=400, detail="vehiculo_id inválido")
        
        combustible_data = {
            "repostado": True,
            "litros": combustible_update.litros,
            "vehiculo_id": combustible_update.vehiculo_id,
            "vehiculo_matricula": vehiculo.get("matricula", ""),
            "km_vehiculo": combustible_update.km_vehiculo,
            "timestamp": get_spain_now(),
            "registrado_por_user_id": str(current_user["_id"])
        }
    else:
        combustible_data = {
            "repostado": False,
            "litros": None,
            "vehiculo_id": None,
            "vehiculo_matricula": None,
            "km_vehiculo": None,
            "timestamp": get_spain_now(),
            "registrado_por_user_id": str(current_user["_id"])
        }
    
    await db.turnos.update_one(
        {"_id": oid, **org_filter},
        {"$set": {"combustible": combustible_data}}
    )
    
    updated_turno = await db.turnos.find_one({"_id": oid, **org_filter})
    
    # Calcular totales
    servicios = await db.services.find({"turno_id": turno_id, **org_filter}).to_list(1000)
    total_clientes = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "empresa")
    total_particulares = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "particular")
    total_km = sum(s.get("kilometros") or 0 for s in servicios)
    
    return TurnoResponse(
        id=turno_id,
        **{k: v for k, v in updated_turno.items() if k != "_id"},
        total_importe_clientes=total_clientes,
        total_importe_particulares=total_particulares,
        total_kilometros=total_km,
        cantidad_servicios=len(servicios)
    )

# (F) COMBUSTIBLE: Estadísticas de combustible
@api_router.get("/turnos/combustible/estadisticas")
async def get_combustible_estadisticas(
    current_user: dict = Depends(get_current_admin),
    from_date: Optional[str] = Query(None, alias="from", description="Fecha inicio (YYYY-MM-DD o dd/mm/yyyy)"),
    to_date: Optional[str] = Query(None, alias="to", description="Fecha fin (YYYY-MM-DD o dd/mm/yyyy)"),
    group: str = Query("day", description="Agrupación: day|week|month")
):
    """
    Estadísticas de combustible (solo admin).
    Respuesta: litros_total, repostajes_total, litros_por_vehiculo, serie por período.
    """
    org_filter = await get_org_filter(current_user)
    
    # Query base: turnos con repostaje
    query = {**org_filter, "combustible.repostado": True}
    
    # Filtro por fechas (formato flexible)
    if from_date:
        # Convertir a formato dd/mm/yyyy si viene en YYYY-MM-DD
        if "-" in from_date:
            parts = from_date.split("-")
            from_date = f"{parts[2]}/{parts[1]}/{parts[0]}"
        query["fecha_inicio"] = {"$gte": from_date}
    if to_date:
        if "-" in to_date:
            parts = to_date.split("-")
            to_date = f"{parts[2]}/{parts[1]}/{parts[0]}"
        if "fecha_inicio" in query:
            query["fecha_inicio"]["$lte"] = to_date
        else:
            query["fecha_inicio"] = {"$lte": to_date}
    
    turnos = await db.turnos.find(query).to_list(10000)
    
    # Calcular estadísticas
    litros_total = 0
    repostajes_total = 0
    litros_por_vehiculo = {}
    serie_por_periodo = {}
    
    for turno in turnos:
        combustible = turno.get("combustible", {})
        if combustible.get("repostado"):
            litros = combustible.get("litros", 0) or 0
            vehiculo_id = combustible.get("vehiculo_id", "")
            vehiculo_matricula = combustible.get("vehiculo_matricula", "Desconocido")
            fecha = turno.get("fecha_inicio", "")
            
            litros_total += litros
            repostajes_total += 1
            
            # Por vehículo
            if vehiculo_id:
                if vehiculo_id not in litros_por_vehiculo:
                    litros_por_vehiculo[vehiculo_id] = {
                        "vehiculo_id": vehiculo_id,
                        "vehiculo_matricula": vehiculo_matricula,
                        "litros": 0,
                        "repostajes": 0
                    }
                litros_por_vehiculo[vehiculo_id]["litros"] += litros
                litros_por_vehiculo[vehiculo_id]["repostajes"] += 1
            
            # Por período (simplificado: por día/semana/mes)
            if fecha:
                # fecha en formato dd/mm/yyyy
                try:
                    parts = fecha.split("/")
                    if group == "day":
                        periodo = fecha
                    elif group == "week":
                        # Semana del año (simplificado)
                        from datetime import datetime as dt
                        d = dt(int(parts[2]), int(parts[1]), int(parts[0]))
                        periodo = f"{d.isocalendar()[0]}-W{d.isocalendar()[1]:02d}"
                    else:  # month
                        periodo = f"{parts[1]}/{parts[2]}"
                    
                    if periodo not in serie_por_periodo:
                        serie_por_periodo[periodo] = {"periodo": periodo, "litros": 0, "repostajes": 0}
                    serie_por_periodo[periodo]["litros"] += litros
                    serie_por_periodo[periodo]["repostajes"] += 1
                except Exception:
                    pass
    
    return {
        "litros_total": round(litros_total, 2),
        "repostajes_total": repostajes_total,
        "litros_por_vehiculo": list(litros_por_vehiculo.values()),
        "serie": sorted(serie_por_periodo.values(), key=lambda x: x["periodo"])
    }

# Exportación de Turnos
@api_router.get("/turnos/export/csv")
async def export_turnos_csv(
    current_user: dict = Depends(get_current_admin),
    taxista_id: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    cerrado: Optional[bool] = Query(None),
    liquidado: Optional[bool] = Query(None)
):
    # SEGURIDAD: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    query = {**org_filter}
    
    # ROBUSTEZ: Si no hay filtros de fecha, limitar a últimos 31 días
    applied_default_limit = False
    if not fecha_inicio and not fecha_fin and not taxista_id:
        from datetime import timedelta
        default_start = (datetime.utcnow() - timedelta(days=31)).strftime("%d/%m/%Y")
        fecha_inicio = default_start
        applied_default_limit = True
        logger.info(f"Export turnos sin filtros: aplicando límite automático desde {default_start}")
    
    # SEGURIDAD: Validar que taxista_id pertenece a la organización
    if taxista_id:
        try:
            taxista_oid = ObjectId(taxista_id)
        except Exception:
            raise HTTPException(status_code=400, detail="taxista_id inválido")
        taxista = await db.users.find_one({"_id": taxista_oid, "role": "taxista", **org_filter})
        if not taxista:
            raise HTTPException(status_code=400, detail="El taxista no existe o no pertenece a esta organización")
        query["taxista_id"] = taxista_id
    
    if fecha_inicio:
        query["fecha_inicio"] = {"$gte": fecha_inicio}
    if fecha_fin:
        if "fecha_inicio" in query:
            query["fecha_inicio"]["$lte"] = fecha_fin
        else:
            query["fecha_inicio"] = {"$lte": fecha_fin}
    if cerrado is not None:
        query["cerrado"] = cerrado
    if liquidado is not None:
        query["liquidado"] = liquidado
    
    turnos = await db.turnos.find(query).sort("fecha_inicio", -1).to_list(10000)
    
    # Usar helper function con org_filter para evitar contaminación de servicios
    turnos_con_totales = await get_turnos_with_servicios(turnos, org_filter=org_filter, include_servicios_detail=True)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header principal de turnos (incluyendo campos de combustible PR1)
    writer.writerow([
        "Tipo", "Taxista", "Vehículo", "Fecha Inicio", "Hora Inicio", "KM Inicio",
        "Fecha Fin", "Hora Fin", "KM Fin", "Total KM",
        "N° Servicios", "Total Clientes (€)", "Total Particulares (€)", "Total (€)",
        "Cerrado", "Liquidado",
        "Comb.Repostado", "Comb.Litros", "Comb.Vehículo", "Comb.KM", "Comb.Fecha",
        "# SERVICIO: Fecha", "Hora", "Origen", "Destino", "Tipo", "Empresa", 
        "Importe", "Importe Espera", "Total", "KM"
    ])
    
    for turno in turnos_con_totales:
        total_km_turno = turno.get("km_fin", 0) - turno["km_inicio"] if turno.get("km_fin") else 0
        total_importe = turno["total_clientes"] + turno["total_particulares"]
        
        # Campos de combustible
        combustible = turno.get("combustible", {}) or {}
        comb_repostado = "Sí" if combustible.get("repostado") else "No"
        comb_litros = combustible.get("litros", "") if combustible.get("repostado") else ""
        comb_vehiculo = combustible.get("vehiculo_matricula", "") if combustible.get("repostado") else ""
        comb_km = combustible.get("km_vehiculo", "") if combustible.get("repostado") else ""
        comb_fecha = combustible.get("timestamp", "").strftime("%d/%m/%Y %H:%M") if combustible.get("timestamp") else ""
        
        # Fila resumen del turno
        writer.writerow([
            "TURNO",
            turno["taxista_nombre"],
            turno["vehiculo_matricula"],
            turno["fecha_inicio"],
            turno["hora_inicio"],
            turno["km_inicio"],
            turno.get("fecha_fin", ""),
            turno.get("hora_fin", ""),
            turno.get("km_fin", ""),
            total_km_turno,
            turno["cantidad_servicios"],
            f"{turno['total_clientes']:.2f}",
            f"{turno['total_particulares']:.2f}",
            f"{total_importe:.2f}",
            "Sí" if turno.get("cerrado") else "No",
            "Sí" if turno.get("liquidado") else "No",
            comb_repostado, comb_litros, comb_vehiculo, comb_km, comb_fecha,
            "", "", "", "", "", "", "", "", "", ""
        ])
        
        # Filas de servicios del turno
        servicios = turno.get("servicios", [])
        for idx, servicio in enumerate(servicios, 1):
            importe = servicio.get("importe", 0)
            importe_espera = servicio.get("importe_espera", 0)
            importe_total = servicio.get("importe_total", importe + importe_espera)
            empresa_nombre = servicio.get("empresa_nombre", "")
            
            writer.writerow([
                "SERVICIO",
                "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                "", "", "", "", "",
                f"#{idx}: {servicio.get('fecha', '')}",
                servicio.get("hora", ""),
                servicio.get("origen", ""),
                servicio.get("destino", ""),
                servicio.get("tipo", ""),
                empresa_nombre if servicio.get("tipo") == "empresa" else "",
                f"{importe:.2f}",
                f"{importe_espera:.2f}",
                f"{importe_total:.2f}",
                servicio.get("kilometros", "") if servicio.get("kilometros") is not None else ""
            ])
        
        # Fila vacía para separar turnos
        writer.writerow([])
    
    output.seek(0)
    headers = {"Content-Disposition": "attachment; filename=turnos_detallado.csv"}
    if applied_default_limit:
        headers["X-Export-Default-Date-Range"] = "31d"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers=headers
    )

@api_router.get("/turnos/export/excel")
async def export_turnos_excel(
    current_user: dict = Depends(get_current_admin),
    taxista_id: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    cerrado: Optional[bool] = Query(None),
    liquidado: Optional[bool] = Query(None)
):
    # SEGURIDAD: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    query = {**org_filter}
    
    # ROBUSTEZ: Si no hay filtros de fecha, limitar a últimos 31 días
    applied_default_limit = False
    if not fecha_inicio and not fecha_fin and not taxista_id:
        from datetime import timedelta
        default_start = (datetime.utcnow() - timedelta(days=31)).strftime("%d/%m/%Y")
        fecha_inicio = default_start
        applied_default_limit = True
        logger.info(f"Export turnos Excel sin filtros: aplicando límite automático desde {default_start}")
    
    # SEGURIDAD: Validar que taxista_id pertenece a la organización
    if taxista_id:
        try:
            taxista_oid = ObjectId(taxista_id)
        except Exception:
            raise HTTPException(status_code=400, detail="taxista_id inválido")
        taxista = await db.users.find_one({"_id": taxista_oid, "role": "taxista", **org_filter})
        if not taxista:
            raise HTTPException(status_code=400, detail="El taxista no existe o no pertenece a esta organización")
        query["taxista_id"] = taxista_id
    
    if fecha_inicio:
        query["fecha_inicio"] = {"$gte": fecha_inicio}
    if fecha_fin:
        if "fecha_inicio" in query:
            query["fecha_inicio"]["$lte"] = fecha_fin
        else:
            query["fecha_inicio"] = {"$lte": fecha_fin}
    if cerrado is not None:
        query["cerrado"] = cerrado
    if liquidado is not None:
        query["liquidado"] = liquidado
    
    turnos = await db.turnos.find(query).sort("fecha_inicio", -1).to_list(10000)
    
    # Usar helper function con org_filter para evitar contaminación de servicios
    turnos_con_totales = await get_turnos_with_servicios(turnos, org_filter=org_filter, include_servicios_detail=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Turnos Detallados"
    
    # Header styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    turno_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    servicio_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    headers = [
        "Tipo", "Taxista", "Vehículo", "Fecha Inicio", "Hora Inicio", "KM Inicio",
        "Fecha Fin", "Hora Fin", "KM Fin", "Total KM",
        "N° Servicios", "Total Clientes (€)", "Total Particulares (€)", "Total (€)",
        "Cerrado", "Liquidado",
        "⛽ Repostó", "⛽ Litros", "⛽ Vehículo", "⛽ KM",
        "Servicio #", "Fecha Serv.", "Hora Serv.", "Origen", "Destino", "Tipo Serv.", 
        "Empresa", "Importe", "Imp. Espera", "Total Serv.", "KM Serv."
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    current_row = 2
    for turno in turnos_con_totales:
        total_km_turno = turno.get("km_fin", 0) - turno["km_inicio"] if turno.get("km_fin") else 0
        total_importe = turno["total_clientes"] + turno["total_particulares"]
        
        # Campos de combustible
        combustible = turno.get("combustible", {}) or {}
        comb_repostado = "Sí" if combustible.get("repostado") else "No"
        comb_litros = combustible.get("litros", "") if combustible.get("repostado") else ""
        comb_vehiculo = combustible.get("vehiculo_matricula", "") if combustible.get("repostado") else ""
        comb_km = combustible.get("km_vehiculo", "") if combustible.get("repostado") else ""
        
        # Fila resumen del turno (con fondo amarillo)
        ws.cell(row=current_row, column=1, value="TURNO")
        ws.cell(row=current_row, column=2, value=turno["taxista_nombre"])
        ws.cell(row=current_row, column=3, value=turno["vehiculo_matricula"])
        ws.cell(row=current_row, column=4, value=turno["fecha_inicio"])
        ws.cell(row=current_row, column=5, value=turno["hora_inicio"])
        ws.cell(row=current_row, column=6, value=turno["km_inicio"])
        ws.cell(row=current_row, column=7, value=turno.get("fecha_fin", ""))
        ws.cell(row=current_row, column=8, value=turno.get("hora_fin", ""))
        ws.cell(row=current_row, column=9, value=turno.get("km_fin", ""))
        ws.cell(row=current_row, column=10, value=total_km_turno)
        ws.cell(row=current_row, column=11, value=turno["cantidad_servicios"])
        ws.cell(row=current_row, column=12, value=round(turno["total_clientes"], 2))
        ws.cell(row=current_row, column=13, value=round(turno["total_particulares"], 2))
        ws.cell(row=current_row, column=14, value=round(total_importe, 2))
        ws.cell(row=current_row, column=15, value="Sí" if turno.get("cerrado") else "No")
        ws.cell(row=current_row, column=16, value="Sí" if turno.get("liquidado") else "No")
        # Columnas de combustible
        ws.cell(row=current_row, column=17, value=comb_repostado)
        ws.cell(row=current_row, column=18, value=comb_litros)
        ws.cell(row=current_row, column=19, value=comb_vehiculo)
        ws.cell(row=current_row, column=20, value=comb_km)
        
        # Aplicar fondo amarillo a la fila del turno
        for col in range(1, 32):
            ws.cell(row=current_row, column=col).fill = turno_fill
        
        current_row += 1
        
        # Filas de servicios del turno (con fondo gris claro)
        servicios = turno.get("servicios", [])
        for idx, servicio in enumerate(servicios, 1):
            importe = servicio.get("importe", 0)
            importe_espera = servicio.get("importe_espera", 0)
            importe_total = servicio.get("importe_total", importe + importe_espera)
            empresa_nombre = servicio.get("empresa_nombre", "")
            
            ws.cell(row=current_row, column=1, value="SERVICIO")
            ws.cell(row=current_row, column=21, value=idx)
            ws.cell(row=current_row, column=22, value=servicio.get("fecha", ""))
            ws.cell(row=current_row, column=23, value=servicio.get("hora", ""))
            ws.cell(row=current_row, column=24, value=servicio.get("origen", ""))
            ws.cell(row=current_row, column=25, value=servicio.get("destino", ""))
            ws.cell(row=current_row, column=26, value=servicio.get("tipo", ""))
            ws.cell(row=current_row, column=27, value=empresa_nombre if servicio.get("tipo") == "empresa" else "")
            ws.cell(row=current_row, column=28, value=round(importe, 2))
            ws.cell(row=current_row, column=29, value=round(importe_espera, 2))
            ws.cell(row=current_row, column=30, value=round(importe_total, 2))
            ws.cell(row=current_row, column=31, value=servicio.get("kilometros", 0) if servicio.get("kilometros") is not None else 0)
            
            # Aplicar fondo gris claro a la fila del servicio
            for col in range(1, 32):
                ws.cell(row=current_row, column=col).fill = servicio_fill
            
            current_row += 1
        
        # Fila vacía para separar turnos
        current_row += 1
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min((max_length + 2), 50)  # Máximo 50 caracteres de ancho
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {"Content-Disposition": "attachment; filename=turnos_detallado.xlsx"}
    if applied_default_limit:
        headers["X-Export-Default-Date-Range"] = "31d"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@api_router.get("/turnos/export/pdf")
async def export_turnos_pdf(
    current_user: dict = Depends(get_current_admin),
    taxista_id: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    cerrado: Optional[bool] = Query(None),
    liquidado: Optional[bool] = Query(None)
):
    # SEGURIDAD: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    query = {**org_filter}
    
    # ROBUSTEZ: Si no hay filtros de fecha, limitar a últimos 31 días
    applied_default_limit = False
    if not fecha_inicio and not fecha_fin and not taxista_id:
        from datetime import timedelta
        default_start = (datetime.utcnow() - timedelta(days=31)).strftime("%d/%m/%Y")
        fecha_inicio = default_start
        applied_default_limit = True
        logger.info(f"Export turnos PDF sin filtros: aplicando límite automático desde {default_start}")
    
    # SEGURIDAD: Validar que taxista_id pertenece a la organización
    if taxista_id:
        try:
            taxista_oid = ObjectId(taxista_id)
        except Exception:
            raise HTTPException(status_code=400, detail="taxista_id inválido")
        taxista = await db.users.find_one({"_id": taxista_oid, "role": "taxista", **org_filter})
        if not taxista:
            raise HTTPException(status_code=400, detail="El taxista no existe o no pertenece a esta organización")
        query["taxista_id"] = taxista_id
    
    if fecha_inicio:
        query["fecha_inicio"] = {"$gte": fecha_inicio}
    if fecha_fin:
        if "fecha_inicio" in query:
            query["fecha_inicio"]["$lte"] = fecha_fin
        else:
            query["fecha_inicio"] = {"$lte": fecha_fin}
    if cerrado is not None:
        query["cerrado"] = cerrado
    if liquidado is not None:
        query["liquidado"] = liquidado
    
    turnos = await db.turnos.find(query).sort("fecha_inicio", -1).to_list(10000)
    
    # Usar helper function con org_filter para evitar contaminación de servicios
    turnos_con_totales = await get_turnos_with_servicios(turnos, org_filter=org_filter, include_servicios_detail=True)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("<b>Turnos Detallados - TaxiFast</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Procesar cada turno con sus servicios
    for turno_idx, turno in enumerate(turnos_con_totales):
        # Título del turno
        turno_title = Paragraph(
            f"<b>Turno {turno_idx + 1}: {turno['taxista_nombre']} - {turno['vehiculo_matricula']}</b>",
            styles['Heading2']
        )
        elements.append(turno_title)
        elements.append(Spacer(1, 0.1*inch))
        
        # Información del turno
        estado = []
        if turno.get("cerrado"):
            estado.append("Cerrado")
        else:
            estado.append("Activo")
        if turno.get("liquidado"):
            estado.append("Liquidado")
        
        total_km_turno = turno.get("km_fin", 0) - turno["km_inicio"] if turno.get("km_fin") else 0
        total_importe = turno["total_clientes"] + turno["total_particulares"]
        
        info_turno = [
            ["Fecha Inicio:", f"{turno['fecha_inicio']} {turno['hora_inicio']}", 
             "Fecha Fin:", f"{turno.get('fecha_fin', 'N/A')} {turno.get('hora_fin', '')}" if turno.get('fecha_fin') else "En curso"],
            ["KM Inicio:", str(turno["km_inicio"]), 
             "KM Fin:", str(turno.get("km_fin", "N/A"))],
            ["Total KM:", str(total_km_turno), 
             "N° Servicios:", str(turno["cantidad_servicios"])],
            ["Total Clientes:", f"{turno['total_clientes']:.2f}€", 
             "Total Particulares:", f"{turno['total_particulares']:.2f}€"],
            ["Total General:", f"{total_importe:.2f}€", 
             "Estado:", " / ".join(estado)]
        ]
        
        # Añadir fila de combustible si repostó
        combustible = turno.get("combustible", {}) or {}
        if combustible.get("repostado"):
            comb_litros = combustible.get("litros", "N/A")
            comb_vehiculo = combustible.get("vehiculo_matricula", "N/A")
            comb_km = combustible.get("km_vehiculo", "N/A")
            info_turno.append([
                "⛽ Repostaje:", f"{comb_litros} L",
                "Vehículo/KM:", f"{comb_vehiculo} / {comb_km} km"
            ])
        
        info_table = Table(info_turno, colWidths=[2*inch, 2*inch, 2*inch, 2*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E7E6E6')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#E7E6E6')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 0.15*inch))
        
        # Tabla de servicios del turno
        servicios = turno.get("servicios", [])
        if servicios:
            servicios_title = Paragraph("<b>Servicios:</b>", styles['Heading3'])
            elements.append(servicios_title)
            elements.append(Spacer(1, 0.05*inch))
            
            servicios_data = [["#", "Fecha", "Hora", "Origen", "Destino", "Tipo", "Importe", "KM"]]
            
            for idx, servicio in enumerate(servicios, 1):
                importe_total = servicio.get("importe_total", servicio.get("importe", 0) + servicio.get("importe_espera", 0))
                origen = servicio.get("origen", "")[:15]
                destino = servicio.get("destino", "")[:15]
                
                servicios_data.append([
                    str(idx),
                    servicio.get("fecha", ""),
                    servicio.get("hora", ""),
                    origen,
                    destino,
                    servicio.get("tipo", "")[:4].upper(),
                    f"{importe_total:.2f}€",
                    str(servicio.get("kilometros", 0))
                ])
            
            servicios_table = Table(servicios_data, colWidths=[0.3*inch, 0.9*inch, 0.7*inch, 1.5*inch, 1.5*inch, 0.6*inch, 0.8*inch, 0.5*inch])
            servicios_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(servicios_table)
        else:
            no_servicios = Paragraph("<i>Este turno no tiene servicios registrados</i>", styles['Normal'])
            elements.append(no_servicios)
        
        # Separador entre turnos
        elements.append(Spacer(1, 0.3*inch))
        if turno_idx < len(turnos_con_totales) - 1:
            elements.append(Paragraph("<hr/>", styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
    
    doc.build(elements)
    
    output.seek(0)
    headers = {"Content-Disposition": "attachment; filename=turnos_detallado.pdf"}
    if applied_default_limit:
        headers["X-Export-Default-Date-Range"] = "31d"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers=headers
    )

# Estadísticas de turnos
@api_router.get("/turnos/estadisticas")
async def get_turnos_estadisticas(
    current_user: dict = Depends(get_current_admin),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None)
):
    # SEGURIDAD: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    query = {**org_filter}
    
    if fecha_inicio:
        query["fecha_inicio"] = {"$gte": fecha_inicio}
    if fecha_fin:
        if "fecha_inicio" in query:
            query["fecha_inicio"]["$lte"] = fecha_fin
        else:
            query["fecha_inicio"] = {"$lte": fecha_fin}
    
    turnos = await db.turnos.find(query).to_list(10000)
    
    total_turnos = len(turnos)
    turnos_cerrados = sum(1 for t in turnos if t.get("cerrado"))
    turnos_liquidados = sum(1 for t in turnos if t.get("liquidado"))
    turnos_activos = total_turnos - turnos_cerrados
    
    # Usar helper function con org_filter para evitar contaminación de servicios
    turnos_con_totales = await get_turnos_with_servicios(turnos, org_filter=org_filter)
    
    total_importe = sum(t["total_clientes"] + t["total_particulares"] for t in turnos_con_totales)
    total_km = sum(t["total_km"] for t in turnos_con_totales)
    total_servicios = sum(t["cantidad_servicios"] for t in turnos_con_totales)
    
    return {
        "total_turnos": total_turnos,
        "turnos_activos": turnos_activos,
        "turnos_cerrados": turnos_cerrados,
        "turnos_liquidados": turnos_liquidados,
        "turnos_pendiente_liquidacion": turnos_cerrados - turnos_liquidados,
        "total_importe": round(total_importe, 2),
        "total_kilometros": round(total_km, 2),
        "total_servicios": total_servicios,
        "promedio_importe_por_turno": round(total_importe / total_turnos, 2) if total_turnos > 0 else 0,
        "promedio_servicios_por_turno": round(total_servicios / total_turnos, 2) if total_turnos > 0 else 0
    }

# Reporte diario por taxista
@api_router.get("/reportes/diario")
async def get_reporte_diario(
    fecha: str = Query(..., description="Fecha en formato dd/mm/yyyy"),
    current_user: dict = Depends(get_current_admin)
):
    """
    Obtiene un reporte diario de todos los taxistas con sus totales del día.
    SEGURIDAD P0: Filtrado por organización.
    Entrada: fecha (dd/mm/yyyy)
    Salida: Lista de taxistas con sus totales del día
    """
    
    # SEGURIDAD P0: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    
    # Obtener taxistas de la organización
    taxistas_query = {"role": "taxista", **org_filter}
    taxistas = await db.users.find(taxistas_query).to_list(1000)
    
    # OPTIMIZACIÓN: Batch query - traer servicios de la fecha filtrados por org
    services_query = {"fecha": fecha, **org_filter}
    all_servicios = await db.services.find(services_query).to_list(10000)
    
    # Agrupar servicios por taxista_id en memoria
    servicios_by_taxista = {}
    for servicio in all_servicios:
        taxista_id = servicio["taxista_id"]
        if taxista_id not in servicios_by_taxista:
            servicios_by_taxista[taxista_id] = []
        servicios_by_taxista[taxista_id].append(servicio)
    
    reporte = []
    
    for taxista in taxistas:
        taxista_id = str(taxista["_id"])
        
        # Obtener servicios del taxista desde el diccionario agrupado
        servicios = servicios_by_taxista.get(taxista_id, [])
        
        if len(servicios) == 0:
            continue  # Omitir taxistas sin servicios ese día
        
        # Calcular totales
        total_servicios = len(servicios)
        total_km = sum(s.get("kilometros") or 0 for s in servicios)
        rec_clientes = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "empresa")
        rec_particulares = sum(s.get("importe_total", s.get("importe", 0)) for s in servicios if s.get("tipo") == "particular")
        
        reporte.append({
            "taxista_id": taxista_id,
            "taxista_nombre": taxista.get("nombre", "Sin nombre"),
            "fecha": fecha,
            "n_servicios": total_servicios,
            "km_totales": round(total_km, 2),
            "rec_clientes": round(rec_clientes, 2),
            "rec_particulares": round(rec_particulares, 2),
            "total": round(rec_clientes + rec_particulares, 2)
        })
    
    # Ordenar por total descendente
    reporte.sort(key=lambda x: x["total"], reverse=True)
    
    return reporte

# ==========================================
# SERVICE ENDPOINTS (Multi-tenant)
# ==========================================
@api_router.post("/services", response_model=ServiceResponse)
async def create_service(service: ServiceCreate, current_user: dict = Depends(get_current_user)):
    """Crear servicio - se asigna a la organización del usuario"""
    
    # SEGURIDAD P1: Superadmin no puede crear servicios (evita datos sin tenant)
    if is_superadmin(current_user):
        raise HTTPException(
            status_code=403,
            detail="Superadmin no puede crear servicios. Use una cuenta de taxista."
        )
    
    org_id = get_user_organization_id(current_user)
    is_admin_user = current_user.get("role") == "admin"
    
    # Si no es admin, buscar turno activo y asignar automáticamente
    turno_activo = None
    if not is_admin_user:
        turno_activo = await db.turnos.find_one({
            "taxista_id": str(current_user["_id"]),
            "cerrado": False
        })
        
        if not turno_activo:
            raise HTTPException(
                status_code=400, 
                detail="Debes iniciar un turno antes de registrar servicios"
            )
    
    # INTEGRIDAD: Validar empresa_id y obtener empresa_nombre desde BD
    empresa_validated = None
    if service.empresa_id:
        empresa_query = {"_id": ObjectId(service.empresa_id)}
        if org_id:
            empresa_query["organization_id"] = org_id
        empresa_validated = await db.companies.find_one(empresa_query)
        if not empresa_validated:
            raise HTTPException(
                status_code=400, 
                detail="La empresa especificada no existe o no pertenece a esta organización"
            )
    
    # INTEGRIDAD: Si admin proporciona turno_id, validar que existe y pertenece a la org
    turno_from_payload = None
    if is_admin_user and service.turno_id:
        turno_query = {"_id": ObjectId(service.turno_id)}
        if org_id:
            turno_query["organization_id"] = org_id
        turno_from_payload = await db.turnos.find_one(turno_query)
        if not turno_from_payload:
            raise HTTPException(
                status_code=400, 
                detail="El turno especificado no existe o no pertenece a esta organización"
            )
    
    # (D) METODO DE PAGO: Validar valores permitidos
    if service.metodo_pago and service.metodo_pago not in ("efectivo", "tpv"):
        raise HTTPException(
            status_code=400,
            detail="metodo_pago debe ser 'efectivo' o 'tpv'"
        )
    
    # (E) ORIGEN TAXITUR: Basado en feature flag de la organizacion
    # Obtener features de la organizacion
    org_features = {}
    if org_id:
        org_doc = await db.organizations.find_one({"_id": ObjectId(org_id)})
        if org_doc:
            org_features = org_doc.get("features", {})
    
    # Verificar si la org tiene el feature taxitur_origen activo
    has_taxitur_origen_feature = org_features.get("taxitur_origen", False)
    
    if has_taxitur_origen_feature:
        # Si el feature esta activo, origen_taxitur es obligatorio
        if not service.origen_taxitur:
            raise HTTPException(
                status_code=400,
                detail="origen_taxitur es obligatorio para esta organizacion (debe ser 'parada' o 'lagos')"
            )
        if service.origen_taxitur not in ("parada", "lagos"):
            raise HTTPException(
                status_code=400,
                detail="origen_taxitur debe ser 'parada' o 'lagos'"
            )
    else:
        # Si el feature NO esta activo, ignorar origen_taxitur (forzar a None)
        service.origen_taxitur = None
    
    # (A) VEHÍCULO EN SERVICIO: Validar y determinar si hubo cambio
    # Determinar vehículo por defecto
    turno_ref = turno_from_payload or turno_activo
    vehiculo_default_id = None
    if turno_ref and turno_ref.get("vehiculo_id"):
        vehiculo_default_id = turno_ref.get("vehiculo_id")
    elif current_user.get("vehiculo_id"):
        vehiculo_default_id = current_user.get("vehiculo_id")
    
    # Si se proporciona vehiculo_id, validar que pertenece al tenant
    vehiculo_validated = None
    vehiculo_cambiado = False
    if service.vehiculo_id:
        try:
            vehiculo_query = {"_id": ObjectId(service.vehiculo_id)}
            if org_id:
                vehiculo_query["organization_id"] = org_id
            vehiculo_validated = await db.vehiculos.find_one(vehiculo_query)
            if not vehiculo_validated:
                raise HTTPException(
                    status_code=400,
                    detail="El vehículo especificado no existe o no pertenece a esta organización"
                )
            # Determinar si hubo cambio de vehículo
            vehiculo_cambiado = (service.vehiculo_id != vehiculo_default_id) if vehiculo_default_id else False
        except Exception as e:
            if "vehiculo_id" in str(e):
                raise
            raise HTTPException(status_code=400, detail="vehiculo_id inválido")
    
    # (A) Si cambió de vehículo, km_inicio y km_fin son obligatorios
    if vehiculo_cambiado:
        if service.km_inicio_vehiculo is None or service.km_fin_vehiculo is None:
            raise HTTPException(
                status_code=400,
                detail="Al cambiar de vehículo, km_inicio_vehiculo y km_fin_vehiculo son obligatorios"
            )
        if service.km_inicio_vehiculo < 0:
            raise HTTPException(
                status_code=400,
                detail="km_inicio_vehiculo debe ser >= 0"
            )
        if service.km_fin_vehiculo < service.km_inicio_vehiculo:
            raise HTTPException(
                status_code=400,
                detail="km_fin_vehiculo debe ser >= km_inicio_vehiculo"
            )
    
    service_dict = service.dict()
    service_dict["taxista_id"] = str(current_user["_id"])
    service_dict["taxista_nombre"] = current_user["nombre"]
    service_dict["created_at"] = datetime.utcnow()
    service_dict["synced"] = True
    
    # Calcular service_dt_utc para ordenacion y filtros correctos
    service_dt_utc = parse_spanish_date_to_utc(service_dict["fecha"], service_dict.get("hora", "00:00"))
    if service_dt_utc:
        service_dict["service_dt_utc"] = service_dt_utc
    
    # Añadir campos de vehículo validados
    service_dict["vehiculo_cambiado"] = vehiculo_cambiado
    if vehiculo_validated:
        service_dict["vehiculo_id"] = str(vehiculo_validated["_id"])
        service_dict["vehiculo_matricula"] = vehiculo_validated.get("matricula", "")
    
    # INTEGRIDAD: Usar empresa_nombre desde BD, ignorar lo que venga del cliente
    if empresa_validated:
        service_dict["empresa_id"] = str(empresa_validated["_id"])
        service_dict["empresa_nombre"] = empresa_validated.get("nombre", "")
    elif "empresa_nombre" in service_dict and not service.empresa_id:
        # Si no hay empresa_id válido, limpiar nombre para evitar inconsistencias
        service_dict["empresa_nombre"] = ""
    
    # Multi-tenant: Asignar organization_id
    service_dict["organization_id"] = org_id
    
    # ========================================
    # IDEMPOTENCIA (Paso 5A): Si hay client_uuid, verificar duplicados
    # ========================================
    client_uuid = service_dict.get("client_uuid")
    if client_uuid:
        # Validar formato de client_uuid
        client_uuid = str(client_uuid).strip()
        if len(client_uuid) < 8 or len(client_uuid) > 64:
            raise HTTPException(
                status_code=400,
                detail="client_uuid debe tener entre 8 y 64 caracteres"
            )
        service_dict["client_uuid"] = client_uuid
        
        # Buscar si ya existe un servicio con este client_uuid en la org
        existing_service = await db.services.find_one({
            "organization_id": org_id,
            "client_uuid": client_uuid
        })
        
        if existing_service:
            # Ya existe - devolver el existente (idempotente)
            return ServiceResponse(
                id=str(existing_service["_id"]),
                **{k: v for k, v in existing_service.items() if k != "_id"}
            )
    else:
        # Sin client_uuid: eliminar del dict para que sparse funcione
        if "client_uuid" in service_dict:
            del service_dict["client_uuid"]
    
    # Asignar turno_id: priorizar el validado del payload, luego turno activo
    if turno_from_payload:
        service_dict["turno_id"] = str(turno_from_payload["_id"])
    elif turno_activo:
        service_dict["turno_id"] = str(turno_activo["_id"])
    
    # Calcular importe_total automaticamente
    service_dict["importe_total"] = service_dict["importe"] + service_dict.get("importe_espera", 0)
    
    # Intentar insertar (puede fallar por DuplicateKeyError si hay concurrencia)
    try:
        result = await db.services.insert_one(service_dict)
        created_service = await db.services.find_one({"_id": result.inserted_id})
    except Exception as e:
        # Si falla por DuplicateKeyError (concurrencia), buscar el existente
        if "duplicate key error" in str(e).lower() and client_uuid:
            existing_service = await db.services.find_one({
                "organization_id": org_id,
                "client_uuid": client_uuid
            })
            if existing_service:
                return ServiceResponse(
                    id=str(existing_service["_id"]),
                    **{k: v for k, v in existing_service.items() if k != "_id"}
                )
        # Si no es DuplicateKeyError o no se encuentra, re-lanzar
        raise HTTPException(status_code=500, detail=f"Error al crear servicio: {str(e)}")
    
    return ServiceResponse(
        id=str(created_service["_id"]),
        **{k: v for k, v in created_service.items() if k != "_id"}
    )

@api_router.post("/services/sync")
async def sync_services(service_sync: ServiceSync, current_user: dict = Depends(get_current_user)):
    """
    Sincronizar servicios offline - se asignan a la organización del usuario.
    INTEGRIDAD: Valida todas las referencias (empresa_id, turno_id) antes de insertar.
    """
    
    # SEGURIDAD P1: Superadmin no puede sincronizar servicios (evita datos sin tenant)
    if is_superadmin(current_user):
        raise HTTPException(
            status_code=403,
            detail="Superadmin no puede sincronizar servicios. Use una cuenta de taxista."
        )
    
    created_services = []
    errors = []
    org_id = get_user_organization_id(current_user)
    is_admin_user = current_user.get("role") == "admin"
    
    # Para taxistas, obtener el turno activo una sola vez
    turno_activo = None
    if not is_admin_user:
        turno_activo = await db.turnos.find_one({
            "taxista_id": str(current_user["_id"]),
            "cerrado": False
        })
    
    # Obtener features de la organizacion para validaciones
    org_features = {}
    if org_id:
        org_doc = await db.organizations.find_one({"_id": ObjectId(org_id)})
        if org_doc:
            org_features = org_doc.get("features", {})
    has_taxitur_origen_feature = org_features.get("taxitur_origen", False)
    
    for idx, service in enumerate(service_sync.services):
        try:
            service_dict = service.dict()
            
            # INTEGRIDAD: Validar empresa_id si viene
            if service.empresa_id:
                empresa_query = {"_id": ObjectId(service.empresa_id)}
                if org_id:
                    empresa_query["organization_id"] = org_id
                empresa = await db.companies.find_one(empresa_query)
                if not empresa:
                    errors.append(f"Servicio {idx}: empresa_id inválido o de otra organización")
                    continue
                # Usar nombre desde BD
                service_dict["empresa_id"] = str(empresa["_id"])
                service_dict["empresa_nombre"] = empresa.get("nombre", "")
            
            # INTEGRIDAD: Validar turno_id
            turno_validated = None
            if service.turno_id:
                turno_query = {"_id": ObjectId(service.turno_id)}
                if org_id:
                    turno_query["organization_id"] = org_id
                # Para taxista, además verificar que es su turno
                if not is_admin_user:
                    turno_query["taxista_id"] = str(current_user["_id"])
                turno_validated = await db.turnos.find_one(turno_query)
                if not turno_validated:
                    errors.append(f"Servicio {idx}: turno_id inválido, de otra organización, o no pertenece al taxista")
                    continue
                service_dict["turno_id"] = str(turno_validated["_id"])
            elif not is_admin_user:
                # Para taxista sin turno_id, asignar turno activo
                if not turno_activo:
                    errors.append(f"Servicio {idx}: no hay turno activo para asignar el servicio")
                    continue
                service_dict["turno_id"] = str(turno_activo["_id"])
                turno_validated = turno_activo
            
            # (7) VALIDACIONES PR1 - Mismas que en POST /api/services
            
            # (D) METODO DE PAGO: Validar valores permitidos
            if service.metodo_pago and service.metodo_pago not in ("efectivo", "tpv"):
                errors.append(f"Servicio {idx}: metodo_pago debe ser 'efectivo' o 'tpv'")
                continue
            
            # (E) ORIGEN TAXITUR: Basado en feature flag de la organizacion
            if has_taxitur_origen_feature:
                if not service.origen_taxitur:
                    errors.append(f"Servicio {idx}: origen_taxitur es obligatorio para esta organizacion")
                    continue
                if service.origen_taxitur not in ("parada", "lagos"):
                    errors.append(f"Servicio {idx}: origen_taxitur debe ser 'parada' o 'lagos'")
                    continue
            else:
                # Si el feature NO esta activo, ignorar origen_taxitur
                service.origen_taxitur = None
            
            # (A) VEHICULO EN SERVICIO: Validar y determinar si hubo cambio
            turno_ref = turno_validated or turno_activo
            vehiculo_default_id = None
            if turno_ref and turno_ref.get("vehiculo_id"):
                vehiculo_default_id = turno_ref.get("vehiculo_id")
            elif current_user.get("vehiculo_id"):
                vehiculo_default_id = current_user.get("vehiculo_id")
            
            vehiculo_cambiado = False
            if service.vehiculo_id:
                try:
                    vehiculo_query = {"_id": ObjectId(service.vehiculo_id)}
                    if org_id:
                        vehiculo_query["organization_id"] = org_id
                    vehiculo_validated = await db.vehiculos.find_one(vehiculo_query)
                    if not vehiculo_validated:
                        errors.append(f"Servicio {idx}: vehiculo_id inválido o de otra organización")
                        continue
                    vehiculo_cambiado = (service.vehiculo_id != vehiculo_default_id) if vehiculo_default_id else False
                    service_dict["vehiculo_id"] = str(vehiculo_validated["_id"])
                    service_dict["vehiculo_matricula"] = vehiculo_validated.get("matricula", "")
                except Exception:
                    errors.append(f"Servicio {idx}: vehiculo_id inválido")
                    continue
            
            service_dict["vehiculo_cambiado"] = vehiculo_cambiado
            
            # Si cambió de vehículo, km_inicio y km_fin son obligatorios
            if vehiculo_cambiado:
                if service.km_inicio_vehiculo is None or service.km_fin_vehiculo is None:
                    errors.append(f"Servicio {idx}: al cambiar vehículo, km_inicio_vehiculo y km_fin_vehiculo son obligatorios")
                    continue
                if service.km_inicio_vehiculo < 0:
                    errors.append(f"Servicio {idx}: km_inicio_vehiculo debe ser >= 0")
                    continue
                if service.km_fin_vehiculo < service.km_inicio_vehiculo:
                    errors.append(f"Servicio {idx}: km_fin_vehiculo debe ser >= km_inicio_vehiculo")
                    continue
            
            # Override con datos del usuario actual
            service_dict["taxista_id"] = str(current_user["_id"])
            service_dict["taxista_nombre"] = current_user["nombre"]
            service_dict["created_at"] = datetime.utcnow()
            service_dict["synced"] = True
            service_dict["organization_id"] = org_id
            
            # Calcular importe_total
            service_dict["importe_total"] = service_dict.get("importe", 0) + service_dict.get("importe_espera", 0)
            
            # Calcular service_dt_utc para ordenacion y filtros correctos
            service_dt_utc = parse_spanish_date_to_utc(service_dict["fecha"], service_dict.get("hora", "00:00"))
            if service_dt_utc:
                service_dict["service_dt_utc"] = service_dt_utc
            
            # ========================================
            # IDEMPOTENCIA (Paso 5A): Verificar client_uuid
            # ========================================
            client_uuid = service_dict.get("client_uuid")
            status = "created_no_uuid"  # Default: sin idempotencia
            
            if client_uuid:
                # Validar formato de client_uuid
                client_uuid = str(client_uuid).strip()
                if len(client_uuid) >= 8 and len(client_uuid) <= 64:
                    service_dict["client_uuid"] = client_uuid
                    
                    # Buscar si ya existe
                    existing_service = await db.services.find_one({
                        "organization_id": org_id,
                        "client_uuid": client_uuid
                    })
                    
                    if existing_service:
                        # Ya existe - marcar como existing
                        created_services.append({
                            "client_uuid": client_uuid,
                            "server_id": str(existing_service["_id"]),
                            "status": "existing"
                        })
                        continue  # No insertar
                    
                    status = "created"  # Con idempotencia
                else:
                    # client_uuid invalido, eliminar del dict
                    del service_dict["client_uuid"]
            else:
                # Sin client_uuid: eliminar del dict para sparse
                if "client_uuid" in service_dict:
                    del service_dict["client_uuid"]
            
            # Intentar insertar
            try:
                result = await db.services.insert_one(service_dict)
                created_services.append({
                    "client_uuid": client_uuid if client_uuid and len(str(client_uuid)) >= 8 else None,
                    "server_id": str(result.inserted_id),
                    "status": status
                })
            except Exception as insert_err:
                # Si falla por DuplicateKeyError, buscar el existente
                if "duplicate key error" in str(insert_err).lower() and client_uuid:
                    existing_service = await db.services.find_one({
                        "organization_id": org_id,
                        "client_uuid": client_uuid
                    })
                    if existing_service:
                        created_services.append({
                            "client_uuid": client_uuid,
                            "server_id": str(existing_service["_id"]),
                            "status": "existing"
                        })
                        continue
                # Otro error
                errors.append(f"Servicio {idx}: error al insertar - {str(insert_err)}")
            
        except Exception as e:
            errors.append(f"Servicio {idx}: error inesperado - {str(e)}")
    
    return {
        "message": f"Processed {len(created_services)} services",
        "results": created_services,
        "errors": errors if errors else None
    }

@api_router.get("/services", response_model=List[ServiceResponse])
async def get_services(
    current_user: dict = Depends(get_current_user),
    tipo: Optional[str] = Query(None),
    empresa_id: Optional[str] = Query(None),
    taxista_id: Optional[str] = Query(None),
    turno_id: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    metodo_pago: Optional[str] = Query(None, description="Filtrar por método de pago: efectivo|tpv"),
    origen_taxitur: Optional[str] = Query(None, description="Filtrar por origen Taxitur: parada|lagos"),
    limit: int = Query(1000, le=10000, description="Límite de resultados")
):
    """Listar servicios - filtrado por organización"""
    # Multi-tenant filter
    org_filter = await get_org_filter(current_user)
    query = {**org_filter}
    
    is_admin_or_super = current_user.get("role") in ["admin", "superadmin"]
    
    # If not admin/superadmin, only show own services
    if not is_admin_or_super:
        query["taxista_id"] = str(current_user["_id"])
    else:
        # SEGURIDAD: Validar taxista_id pertenece al scope
        if taxista_id:
            await _get_taxista_or_400(taxista_id, org_filter, db)
            query["taxista_id"] = taxista_id
    
    # SEGURIDAD: Validar empresa_id pertenece al scope
    if empresa_id:
        await _get_company_or_400(empresa_id, org_filter, db)
        query["empresa_id"] = empresa_id
    
    # SEGURIDAD: Validar turno_id pertenece al scope
    if turno_id:
        taxista_check = None if is_admin_or_super else str(current_user["_id"])
        await _get_turno_or_400(turno_id, org_filter, db, taxista_check)
        query["turno_id"] = turno_id
    
    # Apply filters
    if tipo:
        query["tipo"] = tipo
    
    # Filtros por rango de fechas usando service_dt_utc para queries correctos
    if fecha_inicio or fecha_fin:
        start_utc, end_utc = get_date_range_utc(fecha_inicio, fecha_fin) if fecha_inicio and fecha_fin else (None, None)
        
        # Si ambas fechas son validas, usar service_dt_utc
        if start_utc and end_utc:
            query["service_dt_utc"] = {"$gte": start_utc, "$lte": end_utc}
        elif fecha_inicio:
            # Solo fecha inicio - usar fallback con conversion
            start_utc = parse_spanish_date_to_utc(fecha_inicio, "00:00")
            if start_utc:
                query["service_dt_utc"] = {"$gte": start_utc}
            else:
                # Fallback a string comparison (datos antiguos sin migration)
                query["fecha"] = {"$gte": fecha_inicio}
        elif fecha_fin:
            # Solo fecha fin - usar fallback con conversion
            end_utc = parse_spanish_date_to_utc(fecha_fin, "23:59")
            if end_utc:
                end_utc = end_utc.replace(second=59, microsecond=999999)
                query["service_dt_utc"] = {"$lte": end_utc}
            else:
                # Fallback a string comparison (datos antiguos sin migration)
                query["fecha"] = {"$lte": fecha_fin}
    
    # (D) Filtro por método de pago
    if metodo_pago:
        if metodo_pago in ("efectivo", "tpv"):
            query["metodo_pago"] = metodo_pago
    
    # (E) Filtro por origen Taxitur (solo si la org tiene el feature activo)
    if origen_taxitur and origen_taxitur in ("parada", "lagos"):
        org_id = get_user_organization_id(current_user)
        if org_id:
            # Verificar si la org tiene el feature activo
            org_doc = await db.organizations.find_one({"_id": ObjectId(org_id)})
            if org_doc and org_doc.get("features", {}).get("taxitur_origen", False):
                query["origen_taxitur"] = origen_taxitur
    
    # Validar y ajustar límite
    if limit <= 0:
        limit = 1000  # Default
    elif limit > 10000:
        limit = 10000  # Maximum
    
    # Ordenar por service_dt_utc (datetime real) descendente, fallback a created_at
    services = await db.services.find(query).sort([("service_dt_utc", -1), ("created_at", -1)]).limit(limit).to_list(limit)
    return [
        ServiceResponse(
            id=str(service["_id"]),
            **{k: v for k, v in service.items() if k != "_id"}
        )
        for service in services
    ]

@api_router.put("/services/{service_id}", response_model=ServiceResponse)
async def update_service(service_id: str, service: ServiceCreate, current_user: dict = Depends(get_current_user)):
    """
    Actualizar un servicio existente.
    SEGURIDAD: 
    - Whitelist de campos editables (no se puede cambiar organization_id, created_at, synced, taxista_id)
    - Si admin/superadmin cambia turno_id, se valida que el turno sea de la misma organización
    """
    # SEGURIDAD: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    
    # Check if service exists y pertenece a la organización
    existing_service = await db.services.find_one({"_id": ObjectId(service_id), **org_filter})
    if not existing_service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    # Verificar permisos: admin/superadmin pueden editar cualquiera de su org, taxista solo los suyos
    user_role = current_user.get("role")
    is_admin_or_super = user_role in ["admin", "superadmin"]
    
    if not is_admin_or_super and existing_service["taxista_id"] != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Not authorized to update this service")
    
    # SEGURIDAD: Whitelist de campos que se pueden editar
    # Campos NUNCA editables: organization_id, created_at, synced, taxista_id
    EDITABLE_FIELDS = [
        "origen", "destino", "tipo", "empresa_id", "empresa_nombre",
        "importe", "importe_espera", "observaciones", "cobrado", "facturar", 
        "km", "kilometros", "fecha", "hora"
    ]
    
    # Solo admin/superadmin pueden cambiar turno_id
    if is_admin_or_super:
        EDITABLE_FIELDS.append("turno_id")
    
    service_input = service.dict()
    service_dict = {}
    
    # Solo copiar campos de la whitelist
    for field in EDITABLE_FIELDS:
        if field in service_input and service_input[field] is not None:
            service_dict[field] = service_input[field]
    
    # INTEGRIDAD: Validar que empresa_id pertenece a la misma organización y normalizar nombre
    if "empresa_id" in service_dict and service_dict["empresa_id"]:
        org_id = existing_service.get("organization_id")
        empresa_query = {"_id": ObjectId(service_dict["empresa_id"])}
        if org_id:
            empresa_query["organization_id"] = org_id
        empresa = await db.companies.find_one(empresa_query)
        if not empresa:
            raise HTTPException(
                status_code=400, 
                detail="La empresa especificada no existe o no pertenece a esta organización"
            )
        # INTEGRIDAD: Normalizar empresa_nombre desde BD, ignorar lo que venga del cliente
        service_dict["empresa_nombre"] = empresa.get("nombre", "")
    
    # Validar turno_id si se está cambiando (solo admin/superadmin)
    if "turno_id" in service_dict and service_dict["turno_id"] != existing_service.get("turno_id"):
        # Verificar que el nuevo turno existe y pertenece a la misma organización
        new_turno = await db.turnos.find_one({"_id": ObjectId(service_dict["turno_id"]), **org_filter})
        if not new_turno:
            raise HTTPException(status_code=400, detail="El turno especificado no existe o no pertenece a esta organización")
    
    # Calcular importe_total automáticamente
    importe = service_dict.get("importe", existing_service.get("importe", 0))
    importe_espera = service_dict.get("importe_espera", existing_service.get("importe_espera", 0))
    service_dict["importe_total"] = importe + importe_espera
    
    # Recalcular service_dt_utc si se cambia fecha u hora
    fecha_nueva = service_dict.get("fecha", existing_service.get("fecha"))
    hora_nueva = service_dict.get("hora", existing_service.get("hora", "00:00"))
    if "fecha" in service_dict or "hora" in service_dict:
        service_dt_utc = parse_spanish_date_to_utc(fecha_nueva, hora_nueva)
        if service_dt_utc:
            service_dict["service_dt_utc"] = service_dt_utc
    
    result = await db.services.update_one(
        {"_id": ObjectId(service_id), **org_filter},  # Doble check con org_filter
        {"$set": service_dict}
    )
    
    updated_service = await db.services.find_one({"_id": ObjectId(service_id), **org_filter})
    return ServiceResponse(
        id=str(updated_service["_id"]),
        **{k: v for k, v in updated_service.items() if k != "_id"}
    )

@api_router.delete("/services/{service_id}")
async def delete_service(service_id: str, current_user: dict = Depends(get_current_user)):
    """Eliminar un servicio - con filtro de organización"""
    # SEGURIDAD: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    
    # Check if service exists y pertenece a la organización
    existing_service = await db.services.find_one({"_id": ObjectId(service_id), **org_filter})
    if not existing_service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    user_role = current_user.get("role")
    is_admin_or_super = user_role in ["admin", "superadmin"]
    
    if not is_admin_or_super and existing_service["taxista_id"] != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Not authorized to delete this service")
    
    result = await db.services.delete_one({"_id": ObjectId(service_id), **org_filter})
    return {"message": "Service deleted successfully"}

# Export endpoints
@api_router.get("/services/export/csv")
async def export_csv(
    current_user: dict = Depends(get_current_admin),
    tipo: Optional[str] = Query(None),
    empresa_id: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None)
):
    # SEGURIDAD P0: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    query = {**org_filter}
    
    # ROBUSTEZ: Si no hay filtros de fecha ni empresa, limitar a últimos 31 días
    applied_default_limit = False
    if not fecha_inicio and not fecha_fin and not empresa_id:
        from datetime import timedelta
        default_start = (datetime.utcnow() - timedelta(days=31)).strftime("%d/%m/%Y")
        fecha_inicio = default_start
        applied_default_limit = True
        logger.info(f"Export services CSV sin filtros: aplicando límite automático desde {default_start}")
    
    if tipo:
        query["tipo"] = tipo
    if empresa_id:
        # SEGURIDAD: Validar que empresa pertenece a la org
        await get_empresa_or_400(empresa_id, current_user)
        query["empresa_id"] = empresa_id
    if fecha_inicio:
        query["fecha"] = {"$gte": fecha_inicio}
    if fecha_fin:
        if "fecha" in query:
            query["fecha"]["$lte"] = fecha_fin
        else:
            query["fecha"] = {"$lte": fecha_fin}
    
    services = await db.services.find(query).sort("fecha", 1).to_list(10000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Fecha", "Hora", "Taxista", "Origen", "Destino", "Importe (€)", "Importe Espera (€)", "Importe Total (€)", "Kilómetros", "Tipo", "Empresa", "Cobrado", "Facturar", "Método Pago", "Origen Taxitur", "Vehículo ID", "Vehículo Matrícula", "Vehículo Cambiado", "Km Inicio Vehículo", "Km Fin Vehículo"])
    
    for service in services:
        importe = service.get("importe", 0)
        importe_espera = service.get("importe_espera", 0)
        importe_total = service.get("importe_total", importe + importe_espera)
        
        writer.writerow([
            service["fecha"],
            service["hora"],
            service["taxista_nombre"],
            service["origen"],
            service["destino"],
            f"{importe:.2f}",
            f"{importe_espera:.2f}",
            f"{importe_total:.2f}",
            service.get("kilometros", ""),
            service["tipo"],
            service.get("empresa_nombre", ""),
            "Sí" if service.get("cobrado", False) else "No",
            "Sí" if service.get("facturar", False) else "No",
            service.get("metodo_pago", "") or "",
            service.get("origen_taxitur", "") or "",
            service.get("vehiculo_id", "") or "",
            service.get("vehiculo_matricula", "") or "",
            "Sí" if service.get("vehiculo_cambiado", False) else "No",
            service.get("km_inicio_vehiculo", "") if service.get("km_inicio_vehiculo") is not None else "",
            service.get("km_fin_vehiculo", "") if service.get("km_fin_vehiculo") is not None else ""
        ])
    
    output.seek(0)
    headers = {"Content-Disposition": "attachment; filename=servicios.csv"}
    if applied_default_limit:
        headers["X-Export-Default-Date-Range"] = "31d"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers=headers
    )

@api_router.get("/services/export/excel")
async def export_excel(
    current_user: dict = Depends(get_current_admin),
    tipo: Optional[str] = Query(None),
    empresa_id: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None)
):
    # SEGURIDAD P0: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    query = {**org_filter}
    
    # ROBUSTEZ: Si no hay filtros de fecha ni empresa, limitar a últimos 31 días
    applied_default_limit = False
    if not fecha_inicio and not fecha_fin and not empresa_id:
        from datetime import timedelta
        default_start = (datetime.utcnow() - timedelta(days=31)).strftime("%d/%m/%Y")
        fecha_inicio = default_start
        applied_default_limit = True
        logger.info(f"Export services Excel sin filtros: aplicando límite automático desde {default_start}")
    
    if tipo:
        query["tipo"] = tipo
    if empresa_id:
        # SEGURIDAD: Validar que empresa pertenece a la org
        await get_empresa_or_400(empresa_id, current_user)
        query["empresa_id"] = empresa_id
    if fecha_inicio:
        query["fecha"] = {"$gte": fecha_inicio}
    if fecha_fin:
        if "fecha" in query:
            query["fecha"]["$lte"] = fecha_fin
        else:
            query["fecha"] = {"$lte": fecha_fin}
    
    services = await db.services.find(query).sort("fecha", 1).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios"
    
    # Header styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["Fecha", "Hora", "Taxista", "Origen", "Destino", "Importe (€)", "Importe Espera (€)", "Importe Total (€)", "Kilómetros", "Tipo", "Empresa", "Cobrado", "Facturar", "Método Pago", "Origen Taxitur", "Vehículo ID", "Vehículo Matrícula", "Vehículo Cambiado", "Km Inicio Vehículo", "Km Fin Vehículo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_idx, service in enumerate(services, 2):
        importe = service.get("importe", 0)
        importe_espera = service.get("importe_espera", 0)
        importe_total = service.get("importe_total", importe + importe_espera)
        
        ws.cell(row=row_idx, column=1, value=service["fecha"])
        ws.cell(row=row_idx, column=2, value=service["hora"])
        ws.cell(row=row_idx, column=3, value=service["taxista_nombre"])
        ws.cell(row=row_idx, column=4, value=service["origen"])
        ws.cell(row=row_idx, column=5, value=service["destino"])
        ws.cell(row=row_idx, column=6, value=round(importe, 2))
        ws.cell(row=row_idx, column=7, value=round(importe_espera, 2))
        ws.cell(row=row_idx, column=8, value=round(importe_total, 2))
        ws.cell(row=row_idx, column=9, value=service.get("kilometros", ""))
        ws.cell(row=row_idx, column=10, value=service["tipo"])
        ws.cell(row=row_idx, column=11, value=service.get("empresa_nombre", ""))
        ws.cell(row=row_idx, column=12, value="Sí" if service.get("cobrado", False) else "No")
        ws.cell(row=row_idx, column=13, value="Sí" if service.get("facturar", False) else "No")
        # Nuevos campos PR1
        ws.cell(row=row_idx, column=14, value=service.get("metodo_pago", "") or "")
        ws.cell(row=row_idx, column=15, value=service.get("origen_taxitur", "") or "")
        ws.cell(row=row_idx, column=16, value=service.get("vehiculo_id", "") or "")
        ws.cell(row=row_idx, column=17, value=service.get("vehiculo_matricula", "") or "")
        ws.cell(row=row_idx, column=18, value="Sí" if service.get("vehiculo_cambiado", False) else "No")
        ws.cell(row=row_idx, column=19, value=service.get("km_inicio_vehiculo", "") if service.get("km_inicio_vehiculo") is not None else "")
        ws.cell(row=row_idx, column=20, value=service.get("km_fin_vehiculo", "") if service.get("km_fin_vehiculo") is not None else "")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {"Content-Disposition": "attachment; filename=servicios.xlsx"}
    if applied_default_limit:
        headers["X-Export-Default-Date-Range"] = "31d"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

@api_router.get("/services/export/pdf")
async def export_pdf(
    current_user: dict = Depends(get_current_admin),
    tipo: Optional[str] = Query(None),
    empresa_id: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None)
):
    # SEGURIDAD P0: Filtrar por organización
    org_filter = await get_org_filter(current_user)
    query = {**org_filter}
    
    # ROBUSTEZ: Si no hay filtros de fecha ni empresa, limitar a últimos 31 días
    applied_default_limit = False
    
    # ROBUSTEZ: Si no hay filtros de fecha ni empresa, limitar a últimos 31 días
    if not fecha_inicio and not fecha_fin and not empresa_id:
        from datetime import timedelta
        default_start = (datetime.utcnow() - timedelta(days=31)).strftime("%d/%m/%Y")
        fecha_inicio = default_start
        applied_default_limit = True
        logger.info(f"Export services PDF sin filtros: aplicando límite automático desde {default_start}")
    
    if tipo:
        query["tipo"] = tipo
    if empresa_id:
        # SEGURIDAD: Validar que empresa pertenece a la org
        await get_empresa_or_400(empresa_id, current_user)
        query["empresa_id"] = empresa_id
    if fecha_inicio:
        query["fecha"] = {"$gte": fecha_inicio}
    if fecha_fin:
        if "fecha" in query:
            query["fecha"]["$lte"] = fecha_fin
        else:
            query["fecha"] = {"$lte": fecha_fin}
    
    services = await db.services.find(query).sort("fecha", 1).to_list(10000)
    
    output = io.BytesIO()
    # Usar landscape (horizontal) para tener más espacio
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("<b>Servicios de Taxi - TaxiFast</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data con más columnas (incluyendo nuevos campos PR1)
    data = [["Fecha", "Hora", "Taxista", "Origen", "Destino", "Importe", "Total", "KM", "Tipo", "Cobrado", "Pago", "Orig.Tax", "Veh.Cambio"]]
    
    for service in services:
        importe = service.get("importe", 0)
        importe_espera = service.get("importe_espera", 0)
        importe_total = service.get("importe_total", importe + importe_espera)
        
        data.append([
            service["fecha"],
            service["hora"],
            service["taxista_nombre"][:10],
            service["origen"][:10],
            service["destino"][:10],
            f"{importe:.2f}€",
            f"{importe_total:.2f}€",
            service.get("kilometros", "") or "",
            service["tipo"][:3].upper(),
            "Sí" if service.get("cobrado", False) else "No",
            (service.get("metodo_pago", "") or "")[:3].upper(),
            (service.get("origen_taxitur", "") or "")[:4],
            "Sí" if service.get("vehiculo_cambiado", False) else "No"
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    output.seek(0)
    headers = {"Content-Disposition": "attachment; filename=servicios.pdf"}
    if applied_default_limit:
        headers["X-Export-Default-Date-Range"] = "31d"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers=headers
    )

# Config endpoints
@api_router.get("/config", response_model=ConfigResponse)
async def get_config():
    config = await db.config.find_one()
    if not config:
        # Devolver configuración por defecto
        return ConfigResponse(
            id="default",
            nombre_radio_taxi="TaxiFast",
            telefono="",
            web="www.taxifast.com",
            direccion="",
            email="soporte@taxifast.com",
            logo_base64=None,
            updated_at=datetime.utcnow()
        )
    
    return ConfigResponse(
        id=str(config["_id"]),
        nombre_radio_taxi=config.get("nombre_radio_taxi", "TaxiFast"),
        telefono=config.get("telefono", ""),
        web=config.get("web", ""),
        direccion=config.get("direccion", ""),
        email=config.get("email", ""),
        logo_base64=config.get("logo_base64"),
        updated_at=config.get("updated_at", datetime.utcnow())
    )

@api_router.put("/config", response_model=ConfigResponse)
async def update_config(config: ConfigBase, current_user: dict = Depends(get_current_superadmin)):
    """Actualizar configuración global de la plataforma (SOLO SUPERADMIN)"""
    config_dict = config.dict()
    config_dict["updated_at"] = datetime.utcnow()
    
    existing_config = await db.config.find_one()
    
    if existing_config:
        # Actualizar existente
        await db.config.update_one(
            {"_id": existing_config["_id"]},
            {"$set": config_dict}
        )
        config_id = existing_config["_id"]
    else:
        # Crear nuevo
        result = await db.config.insert_one(config_dict)
        config_id = result.inserted_id
    
    updated_config = await db.config.find_one({"_id": config_id})
    
    return ConfigResponse(
        id=str(updated_config["_id"]),
        **{k: v for k, v in updated_config.items() if k != "_id"}
    )

# Endpoint para que superadmin resetee la configuración global a TaxiFast
@api_router.post("/superadmin/reset-config")
async def superadmin_reset_config(current_user: dict = Depends(get_current_superadmin)):
    """Resetear la configuración global de la plataforma a TaxiFast (solo superadmin)"""
    default_config = {
        "nombre_radio_taxi": "TaxiFast",
        "telefono": "",
        "web": "www.taxifast.com",
        "direccion": "",
        "email": "soporte@taxifast.com",
        "logo_base64": None,
        "updated_at": datetime.utcnow()
    }
    
    # Eliminar configuración existente y crear nueva
    await db.config.delete_many({})
    result = await db.config.insert_one(default_config)
    
    return {
        "message": "Configuración global reseteada a TaxiFast",
        "config": {
            "id": str(result.inserted_id),
            "nombre_radio_taxi": "TaxiFast",
            "web": "www.taxifast.com"
        }
    }

# Endpoint para que superadmin actualice la configuración global
@api_router.put("/superadmin/config")
async def superadmin_update_config(config: dict, current_user: dict = Depends(get_current_superadmin)):
    """Actualizar la configuración global de la plataforma (solo superadmin)"""
    update_data = {
        "updated_at": datetime.utcnow()
    }
    
    if config.get("nombre_radio_taxi"):
        update_data["nombre_radio_taxi"] = config["nombre_radio_taxi"]
    if config.get("telefono") is not None:
        update_data["telefono"] = config["telefono"]
    if config.get("web") is not None:
        update_data["web"] = config["web"]
    if config.get("direccion") is not None:
        update_data["direccion"] = config["direccion"]
    if config.get("email") is not None:
        update_data["email"] = config["email"]
    if config.get("logo_base64") is not None:
        update_data["logo_base64"] = config["logo_base64"]
    
    existing_config = await db.config.find_one()
    
    if existing_config:
        await db.config.update_one(
            {"_id": existing_config["_id"]},
            {"$set": update_data}
        )
    else:
        # Crear con valores por defecto + actualizaciones
        default_config = {
            "nombre_radio_taxi": "TaxiFast",
            "telefono": "",
            "web": "www.taxifast.com",
            "direccion": "",
            "email": "soporte@taxifast.com",
            "logo_base64": None,
        }
        default_config.update(update_data)
        await db.config.insert_one(default_config)
    
    updated_config = await db.config.find_one()
    
    return {
        "message": "Configuración actualizada",
        "config": {
            "nombre_radio_taxi": updated_config.get("nombre_radio_taxi"),
            "telefono": updated_config.get("telefono"),
            "web": updated_config.get("web"),
            "direccion": updated_config.get("direccion"),
            "email": updated_config.get("email")
        }
    }

# ========================================
# MIGRACIÓN INCREMENTAL: datetime fields
# ========================================
MIGRATION_BATCH_SIZE = 2000

async def run_datetime_migration():
    """
    Migración incremental e idempotente para backfill de campos datetime.
    Procesa en batches para no bloquear el arranque.
    """
    try:
        # --- Migración de services.service_dt_utc ---
        migration_key = "services_datetime_v1"
        migration_state = await db.migrations.find_one({"_id": migration_key})
        
        if not migration_state:
            migration_state = {"_id": migration_key, "last_id": None, "done": False, "migrated_count": 0}
            await db.migrations.insert_one(migration_state)
        
        if not migration_state.get("done", False):
            # Buscar services sin service_dt_utc
            query = {"service_dt_utc": {"$exists": False}}
            if migration_state.get("last_id"):
                query["_id"] = {"$gt": ObjectId(migration_state["last_id"])}
            
            services_to_migrate = await db.services.find(query).sort("_id", 1).limit(MIGRATION_BATCH_SIZE).to_list(MIGRATION_BATCH_SIZE)
            
            if services_to_migrate:
                migrated = 0
                last_id = None
                for service in services_to_migrate:
                    fecha = service.get("fecha", "")
                    hora = service.get("hora", "00:00")
                    service_dt_utc = parse_spanish_date_to_utc(fecha, hora)
                    
                    if service_dt_utc:
                        await db.services.update_one(
                            {"_id": service["_id"]},
                            {"$set": {"service_dt_utc": service_dt_utc}}
                        )
                        migrated += 1
                    else:
                        # Fecha malformada - loggear y continuar
                        logger.warning(f"[MIGRATION] Service {service['_id']} tiene fecha malformada: '{fecha}' '{hora}'")
                    
                    last_id = str(service["_id"])
                
                # Actualizar estado de migracion
                total_migrated = migration_state.get("migrated_count", 0) + migrated
                await db.migrations.update_one(
                    {"_id": migration_key},
                    {"$set": {"last_id": last_id, "migrated_count": total_migrated}}
                )
                print(f"[MIGRATION] Services: migrados {migrated} documentos (total: {total_migrated})")
            else:
                # No quedan documentos sin migrar
                await db.migrations.update_one(
                    {"_id": migration_key},
                    {"$set": {"done": True}}
                )
                print(f"[MIGRATION] Services: migracion completada ({migration_state.get('migrated_count', 0)} docs)")
        else:
            print(f"[MIGRATION] Services: ya completada")
        
        # --- Migración de turnos.inicio_dt_utc/fin_dt_utc ---
        migration_key_turnos = "turnos_datetime_v1"
        migration_state_turnos = await db.migrations.find_one({"_id": migration_key_turnos})
        
        if not migration_state_turnos:
            migration_state_turnos = {"_id": migration_key_turnos, "last_id": None, "done": False, "migrated_count": 0}
            await db.migrations.insert_one(migration_state_turnos)
        
        if not migration_state_turnos.get("done", False):
            # Buscar turnos sin inicio_dt_utc
            query = {"inicio_dt_utc": {"$exists": False}}
            if migration_state_turnos.get("last_id"):
                query["_id"] = {"$gt": ObjectId(migration_state_turnos["last_id"])}
            
            turnos_to_migrate = await db.turnos.find(query).sort("_id", 1).limit(MIGRATION_BATCH_SIZE).to_list(MIGRATION_BATCH_SIZE)
            
            if turnos_to_migrate:
                migrated = 0
                last_id = None
                for turno in turnos_to_migrate:
                    update_fields = {}
                    
                    # inicio_dt_utc
                    fecha_inicio = turno.get("fecha_inicio", "")
                    hora_inicio = turno.get("hora_inicio", "00:00")
                    inicio_dt_utc = parse_spanish_date_to_utc(fecha_inicio, hora_inicio)
                    if inicio_dt_utc:
                        update_fields["inicio_dt_utc"] = inicio_dt_utc
                    
                    # fin_dt_utc (solo si tiene fecha_fin)
                    fecha_fin = turno.get("fecha_fin")
                    hora_fin = turno.get("hora_fin", "00:00")
                    if fecha_fin:
                        fin_dt_utc = parse_spanish_date_to_utc(fecha_fin, hora_fin)
                        if fin_dt_utc:
                            update_fields["fin_dt_utc"] = fin_dt_utc
                    
                    if update_fields:
                        await db.turnos.update_one(
                            {"_id": turno["_id"]},
                            {"$set": update_fields}
                        )
                        migrated += 1
                    else:
                        logger.warning(f"[MIGRATION] Turno {turno['_id']} tiene fechas malformadas")
                    
                    last_id = str(turno["_id"])
                
                # Actualizar estado de migracion
                total_migrated = migration_state_turnos.get("migrated_count", 0) + migrated
                await db.migrations.update_one(
                    {"_id": migration_key_turnos},
                    {"$set": {"last_id": last_id, "migrated_count": total_migrated}}
                )
                print(f"[MIGRATION] Turnos: migrados {migrated} documentos (total: {total_migrated})")
            else:
                # No quedan documentos sin migrar
                await db.migrations.update_one(
                    {"_id": migration_key_turnos},
                    {"$set": {"done": True}}
                )
                print(f"[MIGRATION] Turnos: migracion completada ({migration_state_turnos.get('migrated_count', 0)} docs)")
        else:
            print(f"[MIGRATION] Turnos: ya completada")
            
    except Exception as e:
        logger.error(f"[MIGRATION] Error en migracion datetime: {e}")
        # No crashear el startup por errores de migración

# Initialize default admin user and config
@app.on_event("startup")
async def startup_event():
    # ========================================
    # MIGRACIÓN DE ÍNDICES MULTI-TENANT
    # ========================================
    print("[STARTUP] Migrando indices para multi-tenant...")
    
    # --- PRE-LIMPIEZA: Remover client_uuid null/None para que sparse funcione ---
    try:
        # Buscar documentos con client_uuid null o que exista pero sea None
        cleanup_result = await db.services.update_many(
            {"$or": [
                {"client_uuid": None},
                {"client_uuid": {"$type": "null"}},
                {"client_uuid": ""}
            ]},
            {"$unset": {"client_uuid": ""}}
        )
        if cleanup_result.modified_count > 0:
            print(f"[STARTUP] Limpiados {cleanup_result.modified_count} servicios con client_uuid null/vacio")
        else:
            print("[STARTUP] No hay servicios con client_uuid null/vacio para limpiar")
    except Exception as cleanup_err:
        print(f"[STARTUP] Error limpieza client_uuid: {cleanup_err}")
    
    # --- PRE-LIMPIEZA: Eliminar cualquier índice con client_uuid (parcial o completo) ---
    try:
        existing_indexes = await db.services.index_information()
        for idx_name, idx_info in existing_indexes.items():
            if idx_name == "_id_":
                continue
            keys = idx_info.get("key", [])
            # Si el índice tiene client_uuid, eliminarlo
            if any("client_uuid" in str(k) for k in keys):
                print(f"[STARTUP] Eliminando indice existente con client_uuid: {idx_name}")
                await db.services.drop_index(idx_name)
                print(f"[STARTUP] Indice {idx_name} eliminado")
    except Exception as drop_err:
        print(f"[STARTUP] Info: Error eliminando indices client_uuid: {drop_err}")
    
    # --- Migrar índice de vehiculos.matricula ---
    try:
        vehiculos_indexes = await db.vehiculos.index_information()
        # Buscar y eliminar índice único global sobre matricula
        for idx_name, idx_info in vehiculos_indexes.items():
            if idx_name == "_id_":
                continue
            keys = idx_info.get("key", [])
            is_unique = idx_info.get("unique", False)
            # Detectar índice único global: key=[("matricula", 1)] y unique=True
            if is_unique and keys == [("matricula", 1)]:
                print(f"[STARTUP] Eliminando indice unico global de vehiculos: {idx_name}")
                await db.vehiculos.drop_index(idx_name)
                print(f"[STARTUP] Indice {idx_name} eliminado correctamente")
    except Exception as e:
        print(f"[STARTUP] Info: No se pudo verificar indices de vehiculos: {e}")
    
    # --- Migrar índice de companies.numero_cliente ---
    try:
        companies_indexes = await db.companies.index_information()
        # Buscar y eliminar índice único global sobre numero_cliente
        for idx_name, idx_info in companies_indexes.items():
            if idx_name == "_id_":
                continue
            keys = idx_info.get("key", [])
            is_unique = idx_info.get("unique", False)
            # Detectar índice único global: key=[("numero_cliente", 1)] y unique=True
            if is_unique and keys == [("numero_cliente", 1)]:
                print(f"[STARTUP] Eliminando indice unico global de companies: {idx_name}")
                await db.companies.drop_index(idx_name)
                print(f"[STARTUP] Indice {idx_name} eliminado correctamente")
    except Exception as e:
        print(f"[STARTUP] Info: No se pudo verificar indices de companies: {e}")
    
    # ========================================
    # CREAR ÍNDICES DE BASE DE DATOS
    # ========================================
    print("[STARTUP] Creando indices de base de datos...")
    try:
        # Services indexes - Multi-tenant
        await db.services.create_index("turno_id")
        await db.services.create_index("taxista_id")
        await db.services.create_index("fecha")
        await db.services.create_index("tipo")
        await db.services.create_index("organization_id")  # Multi-tenant index
        await db.services.create_index([("fecha", 1), ("taxista_id", 1)])
        await db.services.create_index([("organization_id", 1), ("fecha", 1)])  # Multi-tenant compound
        
        # Turnos indexes - Multi-tenant
        await db.turnos.create_index("taxista_id")
        await db.turnos.create_index("cerrado")
        await db.turnos.create_index("liquidado")
        await db.turnos.create_index("fecha_inicio")
        await db.turnos.create_index("organization_id")  # Multi-tenant index
        await db.turnos.create_index([("taxista_id", 1), ("cerrado", 1)])
        await db.turnos.create_index([("organization_id", 1), ("cerrado", 1)])  # Multi-tenant compound
        
        # Users indexes - Multi-tenant
        await db.users.create_index("username", unique=True)
        await db.users.create_index("role")
        await db.users.create_index("organization_id")  # Multi-tenant index
        await db.users.create_index([("organization_id", 1), ("role", 1)])  # Multi-tenant compound
        
        # Vehiculos indexes - Multi-tenant (ÚNICO POR ORGANIZACIÓN)
        await db.vehiculos.create_index("organization_id")  # Multi-tenant index
        await db.vehiculos.create_index(
            [("organization_id", 1), ("matricula", 1)], 
            unique=True, 
            name="ux_org_matricula"
        )
        print("[STARTUP] Indice ux_org_matricula creado (matricula unica por organizacion)")
        
        # Companies indexes - Multi-tenant (ÚNICO POR ORGANIZACIÓN)
        await db.companies.create_index("organization_id")  # Multi-tenant index
        await db.companies.create_index(
            [("organization_id", 1), ("numero_cliente", 1)], 
            unique=True, 
            sparse=True,
            name="ux_org_numero_cliente"
        )
        print("[STARTUP] Indice ux_org_numero_cliente creado (numero_cliente unico por organizacion)")
        
        # Organizations indexes
        await db.organizations.create_index("slug", unique=True)
        await db.organizations.create_index("activa")
        
        # NUEVOS: Índices para datetime fields (filtros por rango de fechas)
        await db.services.create_index([("organization_id", 1), ("service_dt_utc", -1)], name="idx_org_service_dt")
        await db.turnos.create_index([("organization_id", 1), ("inicio_dt_utc", -1)], name="idx_org_inicio_dt")
        print("[STARTUP] Indices datetime creados (service_dt_utc, inicio_dt_utc)")
        
        # NUEVO: Índice para idempotencia con client_uuid (Paso 5A)
        # Este índice puede fallar si hay datos legacy - lo intentamos pero no bloqueamos
        try:
            # Primero eliminar cualquier índice parcialmente creado
            try:
                existing_indexes = await db.services.index_information()
                if "ux_org_client_uuid" in existing_indexes:
                    await db.services.drop_index("ux_org_client_uuid")
                    print("[STARTUP] Indice ux_org_client_uuid eliminado (existia previamente)")
            except:
                pass
            
            await db.services.create_index(
                [("organization_id", 1), ("client_uuid", 1)],
                unique=True,
                name="ux_org_client_uuid",
                partialFilterExpression={"client_uuid": {"$type": "string", "$exists": True, "$ne": ""}}
            )
            print("[STARTUP] Indice ux_org_client_uuid creado (idempotencia)")
        except Exception as idx_err:
            print(f"[STARTUP] Info: Indice client_uuid no creado (datos legacy): {str(idx_err)[:100]}")
            # La idempotencia funcionara sin el indice pero sera mas lenta
        
        print("[STARTUP] Todos los indices creados correctamente")
    except Exception as e:
        print(f"[STARTUP WARNING] Error creando indices: {e}")
        # No fallar si los índices ya existen
    
    # ========================================
    # MIGRACIÓN INCREMENTAL: Backfill datetime fields
    # ========================================
    await run_datetime_migration()
    
    # Compatibilidad hacia atrás: Si existe TAXITUR_ORG_ID, activar feature flag
    # SOLO SI la key no existe aún (primera vez). Si ya existe (True o False),
    # respetar la decisión del superadmin y NO pisar el valor.
    if TAXITUR_ORG_ID:
        try:
            taxitur_org = await db.organizations.find_one({"_id": ObjectId(TAXITUR_ORG_ID)})
            if taxitur_org:
                features = taxitur_org.get("features") or {}
                if "taxitur_origen" not in features:
                    # Primera vez: activar el feature flag
                    await db.organizations.update_one(
                        {"_id": ObjectId(TAXITUR_ORG_ID)},
                        {"$set": {"features.taxitur_origen": True}}
                    )
                    logger.info(f"[STARTUP] Feature 'taxitur_origen' activado por primera vez para org {TAXITUR_ORG_ID}")
                else:
                    logger.info(f"[STARTUP] Feature 'taxitur_origen' = {features['taxitur_origen']} para org {TAXITUR_ORG_ID} (respetado)")
        except Exception as e:
            logger.warning(f"[STARTUP] No se pudo verificar TAXITUR_ORG_ID: {e}")
    
    # Seed de usuarios por defecto SOLO en no-producción y bajo flag explícita
    # En producción, los usuarios deben crearse manualmente por seguridad
    ALLOW_SEED_USERS = os.environ.get("ALLOW_SEED_USERS", "false").lower() == "true"
    
    if ENV == "production":
        logger.info("User seeding disabled in production")
    else:
        if not ALLOW_SEED_USERS:
            logger.info("User seeding disabled (set ALLOW_SEED_USERS=true to enable in non-production)")
        else:
            # Create default SUPERADMIN if not exists
            superadmin = await db.users.find_one({"role": "superadmin"})
            if not superadmin:
                superadmin_data = {
                    "username": "superadmin",
                    "password": get_password_hash("superadmin123"),
                    "nombre": "Super Administrador TaxiFast",
                    "role": "superadmin",
                    "organization_id": None,
                    "created_at": datetime.utcnow()
                }
                await db.users.insert_one(superadmin_data)
                logger.warning("Default superadmin created in NON-PROD (change password immediately!)")

            # Create default admin if not exists (legacy support)
            admin = await db.users.find_one({"username": "admin"})
            if not admin:
                admin_data = {
                    "username": "admin",
                    "password": get_password_hash("admin123"),
                    "nombre": "Administrador",
                    "role": "admin",
                    "organization_id": None,
                    "created_at": datetime.utcnow()
                }
                await db.users.insert_one(admin_data)
                logger.warning("Default admin created in NON-PROD (change password immediately!)")
    
    # Migrate existing admin to have organization_id field (backward compatibility)
    admin = await db.users.find_one({"username": "admin"})
    if admin and "organization_id" not in admin:
        await db.users.update_one(
            {"_id": admin["_id"]},
            {"$set": {"organization_id": None}}
        )
        logger.info("Existing admin user updated with organization_id field")
    
    # Create default config if not exists (global config / legacy)
    config = await db.config.find_one()
    if not config:
        default_config = {
            "nombre_radio_taxi": "TaxiFast",
            "telefono": "900 000 000",
            "web": "www.taxifast.com",
            "direccion": "España",
            "email": "info@taxifast.com",
            "logo_base64": None,
            "updated_at": datetime.utcnow()
        }
        await db.config.insert_one(default_config)
        logger.info("Default config created")
    
    print("[STARTUP] ✅ TaxiFast Multi-tenant SaaS Platform ready!")

# Include router
app.include_router(api_router)

# CORS Configuration - Seguro para producción
# Orígenes permitidos desde variable de entorno o valores por defecto
CORS_ORIGINS_STR = os.environ.get('CORS_ORIGINS', '')
if CORS_ORIGINS_STR:
    # Si hay orígenes configurados, usarlos
    CORS_ORIGINS = [origin.strip() for origin in CORS_ORIGINS_STR.split(',') if origin.strip()]
else:
    # Valores por defecto para desarrollo y producción conocida
    CORS_ORIGINS = [
        "https://servicios-taxi.vercel.app",
        "https://taxitineo.emergent.host",
        "http://localhost:3000",
        "http://localhost:8081",
        "http://localhost:19006",
    ]

# En desarrollo, permitir todos los orígenes para facilitar pruebas
if ENV == 'development':
    CORS_ORIGINS = ["*"]
    ALLOW_CREDENTIALS = False  # No se puede usar credentials con "*"
else:
    ALLOW_CREDENTIALS = True

app.add_middleware(
    CORSMiddleware,
    allow_credentials=ALLOW_CREDENTIALS,
    allow_origins=CORS_ORIGINS,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
    allow_headers=["Authorization", "Content-Type", "Accept", "Origin", "X-Requested-With"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
